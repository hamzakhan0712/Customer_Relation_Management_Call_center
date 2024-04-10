from audioop import avg
import calendar
import csv
import json
import time
import openpyxl

from datetime import datetime, date, timedelta, time
from django import forms
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import Group
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q, Count, CharField, Sum, F, Avg
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils.dateparse import parse_datetime
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView,View
)

from .models import (
    Agent, AgentAttendance, Break, Calendar, Contact, CustomUser, Lead,
    Marketing, Task, TeamLeader, TeamLeaderAttendance, LeadTransfer,
    PaidCustomer, PaidCustomerNotification, Complaint, ComplaintNotification,
    LeadNotification, ContactNotification, TaskNotification, MarketingNotification, CustomFollowUp,
    CalendarNotification, TeamLeaderNotification, LeadTransferNotification, FollowUpHistory
)
import os
import shutil
from datetime import datetime
from django.http import HttpResponse

def backup_database(request):
    try:
        # Define the backup directory as a subdirectory named 'backup' in the root folder.
        backup_dir = os.path.join(os.getcwd(), 'backup')

        today = datetime.now().strftime('%Y-%m-%d')
        backup_path = os.path.join(backup_dir, today)
        os.makedirs(backup_path, exist_ok=True)

        # Replace 'db.sqlite3' with the actual name of your SQLite database file.
        database_filename = 'db.sqlite3'

        # Define the full paths for the source (original) and destination (backup) database files.
        source_db_path = os.path.join(os.getcwd(), database_filename)
        backup_db_path = os.path.join(backup_path, database_filename)

        # Check if the source database file exists before copying.
        if os.path.exists(source_db_path):
            # Create a copy of the SQLite database file in the backup folder.
            shutil.copy2(source_db_path, backup_db_path)

            print(f'Backup completed and saved to {backup_path}')
            return HttpResponse('Backup completed and saved.')
        else:
            return HttpResponse('Source database not found. Backup failed.')

    except Exception as e:
        return HttpResponse(f'Backup failed with error: {str(e)}')



def user_login(request):
    if request.user.is_authenticated:
        # User is already logged in, redirect to dashboard
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Validate input fields
        if not username or not password:
            messages.error(request, 'Please provide both username and password.')
            return redirect('login')

        user = authenticate(request=request, username=username, password=password)

        if user is not None:
            if user.is_superuser:
                user_type = 'admin'
            else:
                user_type = None
                try:
                    team_leader = TeamLeader.objects.get(auth_user=user)
                    user_type = 'team_leader'
                except TeamLeader.DoesNotExist:
                    pass

                if not user_type:
                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user_type = 'agent'
                    except Agent.DoesNotExist:
                        pass

            if not user_type:
                # Invalid user type (neither admin, team_leader, nor agent)
                messages.error(request, 'Invalid user type.')
                return redirect('login')

            login(request=request, user=user)

            try:
                group = Group.objects.get(name=user_type)
                if group not in user.groups.all():
                    user.groups.add(group)

                if user_type == 'admin':
                    role = 1
                elif user_type == 'team_leader':
                    role = 2
                elif user_type == 'agent':
                    role = 3
                else:
                    role = ""

                user.role = role
                user.save()

                return redirect('dashboard')
            except ObjectDoesNotExist:
                pass

        # Invalid login details
        messages.error(request, 'Invalid login details.')
        return redirect('login')
    
    # GET request, render the login template
    return render(request, 'login.html')



@login_required(login_url='login')
def user_logout(request):
    if 'user_id' in request.session:
        del request.session['user_id']

    role = request.user.role

    if role in [2, 3]:  # Roles 2, 3
        current_date = datetime.now().date()

        if role == 2:
            team_leader = TeamLeader.objects.get(auth_user=request.user)
            attendance = TeamLeaderAttendance.objects.filter(
                team_leader=team_leader, date=current_date, logout_time=None).first()

        elif role == 3:
            agent = Agent.objects.get(auth_user=request.user)
            attendance = AgentAttendance.objects.filter(
                agent=agent, date=current_date, logout_time=None).first()

        if attendance:
            attendance.logout_time = datetime.now()
            attendance.save()

    logout(request)
    return redirect('login')

def calculate_percentage(value, total):
    if total == 0:
        return 0
    return round(((value / total) * 100),2)

@login_required
def dashboard(request):
    user = request.user
    role = user.role  # Assuming the user model has a 'role' field indicating the user's role

    if role == 1:

        # Calculate metrics for Lead model
        leads_count = Lead.objects.all().count()
        new_leads_count = Lead.objects.filter(status='new').count()
        lead_conversion_rate = calculate_percentage(Lead.objects.filter(status='converted').count(), leads_count)

        # Calculate metrics for Task model
        tasks_count = Task.objects.all().count()
        task_completion_rate = calculate_percentage(Task.objects.filter(status='completed').count(), tasks_count)

        # Calculate metrics for Complaint model
        complaint_count = Complaint.objects.all().count()
        complaint_resolution_rate = calculate_percentage(Complaint.objects.filter(resolved=True).count(), complaint_count)

        # Calculate metrics for Marketing model
        marketing_count = Marketing.objects.all().count()
        marketing_success_rate = calculate_percentage(Marketing.objects.filter(campaign_status='completed').count(), marketing_count)

        # Calculate metrics for PaidCustomer model
        paid_customer_count = PaidCustomer.objects.all().count()
        completed_paid_customers_count = PaidCustomer.objects.filter(payment_status='completed').count()
        average_payment_amount = PaidCustomer.objects.aggregate(avg_amount=Avg('amount_paid'))['avg_amount']

        if average_payment_amount is not None:
            average_payment_amount_percentage = (average_payment_amount)
        else:
            average_payment_amount_percentage = 0.0


        metrics = [lead_conversion_rate, task_completion_rate, complaint_resolution_rate, marketing_success_rate]
        metrics_count = len(metrics)
        overall_performance_percentage = sum(metrics) / metrics_count if metrics_count > 0 else 0

        overall_performance_percentage = round(overall_performance_percentage, 2)


        pending_paid_customers_count = PaidCustomer.objects.filter(payment_status='pending').count()


        paused_marketing_campaigns_count = Marketing.objects.filter(campaign_status='paused').count()

        # Count of tasks whose status is pending
        pending_tasks_count = Task.objects.filter(status='pending').count()

        # Count of leads whose status is contacted
        contacted_leads_count = Lead.objects.filter(status='contacted').count()

        # Total sales for the current day
        current_day_sales = PaidCustomer.objects.filter(payment_date=date.today()).aggregate(total_sales=Sum('amount_paid'))['total_sales'] or 0

        # Total sales for the current month
        current_month_sales = PaidCustomer.objects.filter(payment_date__year=date.today().year, payment_date__month=date.today().month).aggregate(total_sales=Sum('amount_paid'))['total_sales'] or 0

        # Total sales for the past 6 months
        past_6_months = date.today() - timedelta(days=6*30)
        past_6_months_sales = PaidCustomer.objects.filter(payment_date__gte=past_6_months).aggregate(total_sales=Sum('amount_paid'))['total_sales'] or 0

        # Total sales for the past one year
        past_one_year = date.today() - timedelta(days=365)
        past_one_year_sales = PaidCustomer.objects.filter(payment_date__gte=past_one_year).aggregate(total_sales=Sum('amount_paid'))['total_sales'] or 0

        sales_data = PaidCustomer.objects.all().order_by('payment_date')
        sales_chart_data = [
            {'payment_date': item.payment_date.strftime('%Y-%m-%d'), 'amount_paid': str(item.amount_paid)}
            for item in sales_data
        ]
        total_sales = sum(item.amount_paid for item in sales_data)

        payment_status_counts = {}
        for status_choice, _ in PaidCustomer.PAYMENT_STATUS_CHOICES:
            count = PaidCustomer.objects.filter(payment_status=status_choice).count()
            payment_status_counts[status_choice] = count
        # Prepare data for the chart
        payment_status_labels = []
        payment_status_counts_chart = []

        for status_choice, status_display in PaidCustomer.PAYMENT_STATUS_CHOICES:
            payment_status_labels.append(status_display)
            payment_status_counts_chart.append(payment_status_counts.get(status_choice, 0))

        # JSON data for the chart
        paid_customers_chart_data = {
            'labels': payment_status_labels,
            'counts': payment_status_counts_chart,
        }
        total_paid_customers = PaidCustomer.objects.all().count()

        leads = Lead.objects.all()
        lead_status_counts = dict(leads.values_list('status').annotate(count=Count('status')))
        lead_status_labels = []
        lead_status_counts_chart = []

        for status_choice, status_display in Lead.LEAD_STATUSES:
            lead_status_labels.append(status_display)
            lead_status_counts_chart.append(lead_status_counts.get(status_choice, 0))

        lead_status_chart_data = {
            'labels': lead_status_labels,
            'counts': lead_status_counts_chart,
        }

        new_leads = Lead.objects.filter(status='new').count()

        current_datetime = datetime.now()

        one_day_ago = current_datetime - timedelta(days=1)
        today = current_datetime.date()

        # Filter objects for each model based on their creation date
        team_leaders = TeamLeader.objects.filter(joining_date__gte=one_day_ago, joining_date__lte=current_datetime)
        agents = Agent.objects.filter(joining_date__gte=one_day_ago, joining_date__lte=current_datetime)
        contacts = Contact.objects.filter(date__gte=one_day_ago, date__lte=current_datetime)
        tasks = Task.objects.filter(date__gte=one_day_ago, date__lte=current_datetime)
        complaints = Complaint.objects.filter(created_at__gte=one_day_ago, created_at__lte=current_datetime)
        marketing_campaigns = Marketing.objects.filter(date__gte=one_day_ago, date__lte=current_datetime)
        calendars = Calendar.objects.filter(event_date__gte=one_day_ago, event_date__lte=current_datetime)
        paid_customers = PaidCustomer.objects.filter(payment_date__gte=one_day_ago, payment_date__lte=current_datetime)

        data = {
            'team_leaders': list(team_leaders.values()),
            'agents': list(agents.values()),
            'contacts': list(contacts.values()),
            'tasks': list(tasks.values()), 
            'complaints': list(complaints.values()),
            'marketing_campaigns': list(marketing_campaigns.values()),
            'calendars': list(calendars.values()),
            'paid_customers': list(paid_customers.values()),
        }


        return render(request, 'dashboard_admin.html', {
            'lead_conversion_rate': lead_conversion_rate,
            'task_completion_rate': task_completion_rate,
            'complaint_resolution_rate': complaint_resolution_rate,
            'marketing_success_rate': marketing_success_rate,
            'average_payment_amount': average_payment_amount,
            'average_payment_amount_percentage':average_payment_amount_percentage,
            'overall_performance_percentage': overall_performance_percentage,
            'pending_paid_customers_count': pending_paid_customers_count,
            'paused_marketing_campaigns_count': paused_marketing_campaigns_count,
            'pending_tasks_count': pending_tasks_count,
            'contacted_leads_count': contacted_leads_count,
            'lead_status_chart_data': json.dumps(lead_status_chart_data),
            'paid_customers_chart_data': json.dumps(paid_customers_chart_data),
            'sales_data_json': json.dumps(sales_chart_data),
            'total_sales': total_sales,
            'total_paid_customers': total_paid_customers,
            'new_leads':new_leads,
            'data': data,
            'current_day_sales' : current_day_sales,
            'current_month_sales' : current_month_sales,
            'past_6_months_sales' : past_6_months_sales,
            'past_one_year_sales' : past_one_year_sales,
            
        })

    elif role == 2:
        current_date = datetime.now()
        team_leader = TeamLeader.objects.get(auth_user=user)
        team_leader_attendance = TeamLeaderAttendance.objects.filter(team_leader=team_leader)
        existing_attendance = TeamLeaderAttendance.objects.filter(team_leader=team_leader, date=current_date).exists()
        leaves_taken = team_leader_attendance.filter(status='Absent').count()
        

        total_agents_count = Agent.objects.filter(team_leader=team_leader).count()
        contacted_leads = Lead.objects.filter(status='contacted', assigned_to=team_leader).count()
        pending_paid_customers = PaidCustomer.objects.filter(payment_status='pending',lead__assigned_to=team_leader).count()
        total_paid_customers = PaidCustomer.objects.filter(lead__assigned_to=team_leader).count()

        active_marketing_campaigns = Marketing.objects.filter(campaign_status='active').count()

        agents = Agent.objects.filter(team_leader=team_leader)
        agents_custom_users = [agent.auth_user for agent in agents]

        complaints_count = Complaint.objects.filter(Q(assigned_to=user) | Q(assigned_to__in=agents_custom_users)).count()

        current_month = current_date.month
        current_year = current_date.year

        total_team_leader_sales = PaidCustomer.objects.filter(
            payment_status='completed',
            lead__assigned_to=team_leader,
            payment_date__month=current_month,
            payment_date__year=current_year
        ).aggregate(total_sales=Sum('amount_paid'))['total_sales'] or 0

        # Reset the total sales to zero at the beginning of the next month
        next_month = current_month + 1 if current_month < 12 else 1
        next_year = current_year + 1 if current_month == 12 else current_year

        if current_date.day == 1:
            total_team_leader_sales = 0




        pending_task_count = Task.objects.filter(Q(assigned_to=user) | Q(assigned_to__in=agents_custom_users)).count()

        next_two_days = current_date + timedelta(days=2)
        upcoming_events = Calendar.objects.filter(Q(assigned_to=user) | Q(assigned_to__in=agents_custom_users),event_date__range=[current_date, next_two_days])

        previous_day = current_date - timedelta(days=1)
        previous_day_contacted_leads = Lead.objects.filter(status='contacted', assigned_to=team_leader,date=previous_day).count()
        previous_day_new_customer = PaidCustomer.objects.filter(payment_status='pending', lead__assigned_to=team_leader, payment_date=previous_day).count()
        previous_day_cpmpleted_task = Task.objects.filter(Q(assigned_to=user) | Q(assigned_to__in=agents_custom_users),status='completed',date=previous_day).count()

        yesterday = current_date - timedelta(days=1)

        # Querysets for recent objects within the date range (yesterday and today)
        recent_agents = Agent.objects.filter(joining_date__range=[yesterday, current_date])
        recent_paid_customers = PaidCustomer.objects.filter(payment_date__range=[yesterday, current_date])
        recent_contacts = Contact.objects.filter(date__range=[yesterday, current_date])
        recent_tasks = Task.objects.filter(date__range=[yesterday, current_date])
        recent_complaints = Complaint.objects.filter(created_at__date__range=[yesterday, current_date])
        recent_marketing_campaigns = Marketing.objects.filter(date__range=[yesterday, current_date])
        recent_calendars = Calendar.objects.filter(Q(assigned_to=user) | Q(assigned_to__in=agents_custom_users),event_date__range=[yesterday, current_date])

         # Get the agent attendance records for today
        agent_attendance_today = AgentAttendance.objects.filter(agent__in=agents, date=current_date)
        agent_attendance_yesterday = AgentAttendance.objects.filter(agent__in=agents, date=yesterday)

        # Calculate the counts for agents who are present, late, and absent
        present_count = agent_attendance_today.filter(status='Present').count()
        late_count = agent_attendance_today.filter(status='Late').count()
        absent_count = agent_attendance_yesterday.filter(status='Absent').count()


        team_leader_attendance_today = TeamLeaderAttendance.objects.filter(team_leader=user.teamleader, date=current_date).first()

        # Retrieve the login time from the team leader attendance record
        login_time = team_leader_attendance_today.login_time if team_leader_attendance_today else None

        previous_2_days = current_date - timedelta(days=2)

        # Retrieve the follow-up history meeting the criteria
        recent_leads = Lead.objects.filter(assigned_to=team_leader,date__range=[previous_2_days, current_date])
        
        reminders = LeadReminder.objects.filter(lead__assigned_to=team_leader,reminder_date__date=current_date)

        if existing_attendance:
            return render(request, 'dashboard_team_leader.html', {
                'attendance': team_leader_attendance,
                'leaves_taken': leaves_taken,
                'show_form': False,
                'total_agents_count': total_agents_count,
                'contacted_leads':contacted_leads,
                'pending_paid_customers': pending_paid_customers,
                'active_marketing_campaigns':active_marketing_campaigns,
                'agents': agents_custom_users,
                'complaints_count':complaints_count,
                'total_team_leader_sales':total_team_leader_sales,
                'total_paid_customers':total_paid_customers,
                'pending_task_count':pending_task_count,
                'upcoming_events':upcoming_events,
                'previous_day_contacted_leads':previous_day_contacted_leads,
                'previous_day_new_customer':previous_day_new_customer,
                'previous_day_cpmpleted_task':previous_day_cpmpleted_task,
                'recent_agents': recent_agents,
                'recent_paid_customers': recent_paid_customers,
                'recent_contacts': recent_contacts,
                'recent_tasks': recent_tasks,
                'recent_complaints': recent_complaints,
                'recent_marketing_campaigns': recent_marketing_campaigns,
                'recent_calendars': recent_calendars,
                'present_count': present_count,
                'late_count': late_count,
                'absent_count': absent_count,
                'login_time': login_time, 
                'recent_leads': recent_leads, 
                'reminders':reminders
            })

        if request.method == 'POST':
            login_time = datetime.now()  # Use naive datetime for login_time
            logout_time = None
            current_time = datetime.now().time()
            working_place = request.POST.get('working_place')

            on_leave = request.POST.get('on_leave', False) == 'on'


            if on_leave:
                # If "On Leave" checkbox is checked, set the status to "Absent"
                status = 'Absent'
            else:
                # If "On Leave" checkbox is not checked, determine the status based on the current time
                current_time = datetime.now().time()
                if time(9, 0) <= current_time < time(13, 0):
                    status = 'Present'
                elif time(13, 0) <= current_time < time(18, 0):
                    status = 'Late'
                else: 
                    status = ''

            # Save the attendance entry
            leaves_taken = on_leave


            attendance = TeamLeaderAttendance.objects.create(
                team_leader=team_leader,
                date=current_date,
                status=status,
                working_place=working_place,
                login_time=login_time,
                logout_time=logout_time,
                leaves_taken=leaves_taken,
            )
            attendance.save()

            return redirect('dashboard')
        

        return render(request, 'dashboard_team_leader.html', {
            'attendance': team_leader_attendance,
            'leaves_taken': leaves_taken,
            'show_form': True,
        })

    elif role == 3:
        current_date = datetime.now()
        agent = Agent.objects.get(auth_user=user)
        agent_attendance = AgentAttendance.objects.filter(agent=agent)
        existing_attendance = AgentAttendance.objects.filter(agent=agent, date=current_date).exists()
        leaves_taken = agent_attendance.filter(status='Absent').count()  # Initialize leaves_taken

        agent_attendance_today = AgentAttendance.objects.filter(agent=user.agent, date=current_date).first()
        login_time = agent_attendance_today.login_time if agent_attendance_today else None

        next_two_days = current_date + timedelta(days=1)
        calender_event = Calendar.objects.filter(assigned_to=user,event_date__range=[current_date,next_two_days])

        total_sales_agent = PaidCustomer.objects.filter(payment_status='completed',lead__assigned_to_agents=agent).aggregate(total_sales=Sum('amount_paid'))['total_sales'] or 0
        total_paid_customers = PaidCustomer.objects.filter(lead__assigned_to_agents=agent)
        total_task_count = Task.objects.filter(assigned_to=user,status='in_progress')
        pending_leads =  Lead.objects.filter(status='new', assigned_to_agents=agent)
        pending_customer = PaidCustomer.objects.filter(payment_status='pending',lead__assigned_to_agents=agent)
        pending_task = Task.objects.filter(assigned_to=user,status='pending')
        unresolved_complaints = Complaint.objects.filter(assigned_to=user)

        previous_day = current_date - timedelta(days=1)
        previous_leads = Lead.objects.filter(
            Q(date=previous_day, status='new', assigned_to_agents=agent) |
            Q(date=previous_day, status='contacted', assigned_to_agents=agent)
        )

        recent_complaints = Complaint.objects.filter(assigned_to=user,resolved=False)
        reminders = LeadReminder.objects.filter(lead__assigned_to_agents=agent,reminder_date__date=current_date)
        if existing_attendance:
            return render(request, 'dashboard_agent.html', {
                'attendance': agent_attendance, 
                'leaves_taken': leaves_taken, 
                'show_form': False,
                'login_time': login_time,
                'calender_event':calender_event,
                'total_sales_agent':total_sales_agent,
                'total_paid_customers':total_paid_customers,
                'total_task_count':total_task_count,
                'pending_leads':pending_leads,
                'pending_customer':pending_customer,
                'pending_task':pending_task,
                'unresolved_complaints':unresolved_complaints,
                'previous_leads':previous_leads,
                'recent_complaints':recent_complaints,
                'reminders':reminders

            })
   
        if request.method == 'POST':
            login_time = datetime.now()  # Use naive datetime for login_time
            logout_time = None
            current_time = datetime.now().time()
            working_place = request.POST.get('working_place')

            on_leave = request.POST.get('on_leave', False) == 'on'

            if on_leave:
                # If "On Leave" checkbox is checked, set the status to "Absent"
                status = 'Absent'
            else:
                # If "On Leave" checkbox is not checked, determine the status based on the current time
                current_time = datetime.now().time()
                if time(1, 0) <= current_time < time(12, 0):
                    status = 'Present'
                elif time(12, 0) <= current_time < time(18, 0):
                    status = 'Late'
                else:
                    status = ''


            # Save the attendance entry
            leaves_taken = on_leave

            attendance = AgentAttendance.objects.create(
                agent=agent,
                date=current_date,
                status=status,
                working_place=working_place,
                login_time=login_time,
                logout_time=logout_time,
                leaves_taken=leaves_taken,
            )
            attendance.save()

            return redirect('dashboard')

        return render(request, 'dashboard_agent.html', {'attendance': agent_attendance, 'leaves_taken': leaves_taken, 'show_form': True})

    else:
        return redirect('login')


@login_required
def inbox(request):
    if request.user.role == 2:  # Team Leader
        team_leader = request.user.teamleader

        # Fetch notifications for both from_team_leader and to_team_leader
        notifications = TeamLeaderNotification.objects.filter(
            Q(team_leader=team_leader.id) |
            Q(message__contains=f"to Team Leader {team_leader.first_name} {team_leader.last_name}.") |
            Q(message__contains=f"from Team Leader {team_leader.first_name} {team_leader.last_name}.") |
            Q(message__contains=f"to Team Leader {team_leader.first_name} {team_leader.last_name} (Lead Transfer).")
        ).order_by('-created_at')

    elif request.user.role == 3:  # Agent
        agent = request.user.agent

        # Fetch notifications for both from_agent and to_agent
        notifications = TeamLeaderNotification.objects.filter(
            Q(agent=agent.id) |
            Q(message__contains=f"to Agent {agent.first_name} {agent.last_name}.") |
            Q(message__contains=f"from Agent {agent.first_name} {agent.last_name}.") |
            Q(message__contains=f"to Agent {agent.first_name} {agent.last_name} (Lead Transfer).")
        ).order_by('-created_at')
    else:
        notifications = TeamLeaderNotification.objects.none()

    notification_count = notifications.count()

    if request.user.is_authenticated:
        current_user = request.user
        TeamLeaderNotification.objects.filter(user=current_user).update(seen=True)

    return render(request, 'inbox.html', {
        'notification_count': notification_count,
        'notifications': notifications
    })


@login_required
def agent_reports(request):
    agent_search_query = request.GET.get("agent_search", "")
    agent_date_from_str = request.GET.get("agent_date_from", "")
    agent_date_to_str = request.GET.get("agent_date_to", "")

    if request.user.role == 2:
        team_leader = TeamLeader.objects.get(auth_user=request.user)
        agents = Agent.objects.filter(team_leader=team_leader)
    else:
        agents = Agent.objects.all()

    if agent_search_query:
        agents = agents.filter(
            Q(first_name__icontains=agent_search_query) |
            Q(last_name__icontains=agent_search_query) |
            Q(team_leader__first_name__icontains=agent_search_query) |
            Q(team_leader__last_name__icontains=agent_search_query)
        )

    agent_data = []
    for agent in agents:
        leads = Lead.objects.filter(status='converted',assigned_to_agents=agent)
        paid_customers = PaidCustomer.objects.filter(lead__assigned_to_agents=agent,payment_status='completed')
        working_days = AgentAttendance.objects.filter(
            Q(agent=agent, status='Present') | Q(agent=agent, status='Late')
        )
        leaves_taken = AgentAttendance.objects.filter(agent=agent,status='Absent')
        tasks = Task.objects.filter(assigned_to=agent.auth_user,status='completed')
        complaints = Complaint.objects.filter(assigned_to=agent.auth_user,resolved=True)

        if agent_date_from_str and agent_date_to_str:
            agent_date_from = datetime.strptime(agent_date_from_str, "%Y-%m-%d").date()
            agent_date_to = datetime.strptime(agent_date_to_str, "%Y-%m-%d").date()
            if agent.joining_date >= agent_date_from and agent.joining_date <= agent_date_to:
                agent_data.append({
                    'agent': agent,
                    'leads': leads,
                    'working_days':working_days,
                    'paid_customers': paid_customers,
                    'leaves_taken': leaves_taken,
                    'tasks':tasks,
                    'complaints':complaints
                })
        else:
            agent_data.append({
                'agent': agent,
                'leads': leads,
                'working_days':working_days,
                'paid_customers': paid_customers,
                'leaves_taken': leaves_taken,
                'tasks':tasks,
                'complaints':complaints
            })

    context = {
        'agent_data': agent_data,
        'agent_search_query': agent_search_query,
        'agent_date_from': agent_date_from_str,  
        'agent_date_to': agent_date_to_str,      
    }

    return render(request, 'agent_reports.html', context)


@login_required
def agent_performance_report(request, agent_id):
    agent = get_object_or_404(Agent, id=agent_id)
    leads = Lead.objects.filter(assigned_to_agents=agent)
    paid_customers = PaidCustomer.objects.filter(lead__assigned_to_agents=agent)

    leaves_taken = AgentAttendance.objects.filter(agent=agent,status='Absent')
    tasks = Task.objects.filter(assigned_to=agent.auth_user)
    complaints = Complaint.objects.filter(assigned_to=agent.auth_user)

    context = {
        'agent': agent,
        'leads': leads,
     
        'paid_customers': paid_customers,
        'leaves_taken': leaves_taken,
        'tasks':tasks,
        'complaints':complaints
    }

    return render(request, 'agent_performance_report.html', context)


@login_required
def agent_attendance_report(request, agent_id):
    agent = get_object_or_404(Agent, id=agent_id)
    attendance = AgentAttendance.objects.filter(agent=agent)
    present_attendance = AgentAttendance.objects.filter(agent=agent, status='Present')
    absent_attendance = AgentAttendance.objects.filter(agent=agent, status='Absent')
    late_attendance = AgentAttendance.objects.filter(agent=agent, status='Late')
    total_working_days = present_attendance.count() + late_attendance.count()

    agent_date_from_str = request.GET.get("agent_date_from", "")
    agent_date_to_str = request.GET.get("agent_date_to", "")

    if agent_date_from_str and agent_date_to_str:
        agent_date_from = timezone.datetime.strptime(agent_date_from_str, "%Y-%m-%d")
        agent_date_to = timezone.datetime.strptime(agent_date_to_str, "%Y-%m-%d")
        attendance = attendance.filter(date__range=(agent_date_from, agent_date_to))
        present_attendance = present_attendance.filter(date__range=(agent_date_from, agent_date_to))
        absent_attendance = absent_attendance.filter(date__range=(agent_date_from, agent_date_to))
        late_attendance = late_attendance.filter(date__range=(agent_date_from, agent_date_to))
        total_working_days = present_attendance.count() + late_attendance.count()

    sort_by = request.GET.get("sort")
    if sort_by == "date":
        attendance = attendance.order_by("date")
    elif sort_by == "status":
        attendance = attendance.order_by("status")

    filter_by = request.GET.get("filter")
    if filter_by == "today":
        today = timezone.now().date()
        attendance = attendance.filter(date=today)
    elif filter_by == "this_week":
        start_of_week = timezone.now().date() - timezone.timedelta(days=timezone.now().weekday())
        end_of_week = start_of_week + timezone.timedelta(days=6)
        attendance = attendance.filter(date__range=[start_of_week, end_of_week])
    elif filter_by == "this_month":
        today = timezone.now().date()
        start_of_month = today.replace(day=1)
        end_of_month = start_of_month.replace(day=calendar.monthrange(start_of_month.year, start_of_month.month)[1])
        attendance = attendance.filter(date__range=[start_of_month, end_of_month])
    elif filter_by == "6_months":
        today = timezone.now().date()
        start_of_six_months_ago = today - timezone.timedelta(days=180)
        attendance = attendance.filter(date__gte=start_of_six_months_ago)
    elif filter_by == "this_year":
        today = timezone.now().date()
        start_of_year = today.replace(month=1, day=1)
        end_of_year = today.replace(month=12, day=31)
        attendance = attendance.filter(date__range=[start_of_year, end_of_year])

    context = {
        'agent': agent,
        'attendance': attendance,
        'present_attendance': present_attendance,
        'absent_attendance': absent_attendance,
        'late_attendance': late_attendance,
        'total_working_days': total_working_days,
        'agent_date_from': agent_date_from_str,
        'agent_date_to': agent_date_to_str,
    }

    return render(request, 'agent_attendance_report.html', context)


@login_required
def export_agent_attendance(request, agent_id):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filter_value = request.GET.get("filter", "")
    
    # Fetch the agent data
    agent = get_object_or_404(Agent, id=agent_id)

    # Set the Excel file name
    response['Content-Disposition'] = f'attachment; filename="{agent.first_name}_{agent.last_name}_Agent_AttendanceReport.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the headers for the first table (Agent Information)
    headers_agent = ['Agent Name', 'Phone Number', 'Total Working Days', 'Days Present', 'Days Late', 'Days Absent']
    for col_num, header in enumerate(headers_agent, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Fetch the attendance data for the specific agent
    agent_attendance = AgentAttendance.objects.filter(agent=agent)
    present_attendance = agent_attendance.filter(status='Present')
    absent_attendance = agent_attendance.filter(status='Absent')
    late_attendance = agent_attendance.filter(status='Late')
    total_working_days = present_attendance.count() + late_attendance.count()

    # Write the data for the first table (Agent Information)
    data_agent = [f"{agent.first_name} {agent.last_name}",
                  agent.phone_number,
                  total_working_days,
                  present_attendance.count(),
                  late_attendance.count(),
                  absent_attendance.count()]

    for col_num, value in enumerate(data_agent, 1):
        worksheet.cell(row=2, column=col_num, value=value)

    # Write the headers for the second table (Attendance Data)
    headers_attendance = ['Date', 'Status', 'Working Place', 'Login Time', 'Logout Time']
    for col_num, header in enumerate(headers_attendance, 1):
        worksheet.cell(row=4, column=col_num, value=header)

    # Fetch the attendance data for the specific agent based on filters (if any)
    if filter_value == 'today':
        attendance_data = AgentAttendance.objects.filter(agent=agent, date=date.today())
    elif filter_value == 'this_week':
        start_date = date.today() - timedelta(days=date.today().weekday())
        end_date = start_date + timedelta(days=6)
        attendance_data = AgentAttendance.objects.filter(agent=agent, date__range=[start_date, end_date])
    elif filter_value == 'this_month':
        start_date = date.today().replace(day=1)
        end_date = date.today().replace(day=calendar.monthrange(date.today().year, date.today().month)[1])
        attendance_data = AgentAttendance.objects.filter(agent=agent, date__range=[start_date, end_date])
    elif filter_value == '6_months':
        today = date.today()
        start_date = today - timedelta(days=180)
        end_date = today
        attendance_data = AgentAttendance.objects.filter(agent=agent, date__range=[start_date, end_date])
    elif filter_value == 'this_year':
        start_date = date.today().replace(month=1, day=1)
        end_date = date.today().replace(month=12, day=31)
        attendance_data = AgentAttendance.objects.filter(agent=agent, date__range=[start_date, end_date])
    else:
        attendance_data = AgentAttendance.objects.filter(agent=agent)

    for row_num, attendance in enumerate(attendance_data, 5):
        # Format the date as a string in the desired format (e.g., "YYYY-MM-DD")
        date_str = attendance.date.strftime("%Y-%m-%d")

        # Format login_time and logout_time only if they are not None
        login_time_str = attendance.login_time.strftime("%H:%M:%S") if attendance.login_time else ""
        logout_time_str = attendance.logout_time.strftime("%H:%M:%S") if attendance.logout_time else ""

        # Write the attendance data to the second table
        data_attendance = [date_str,
                           attendance.status,
                           attendance.working_place,
                           login_time_str,
                           logout_time_str,
                           ]

        for col_num, value in enumerate(data_attendance, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(response)
    return response


@login_required
def export_agent_reports(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="agent_report.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the headers for the Excel
    headers = ['Joining Date', 'Agent Full Name', 'Under Team leader', 'Leads Count', 'Paid Customers', 'Task Completed', 'Call Attend', 'Resolved Complaint', 'Leaves Taken']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Fetch all the agents
    if request.user.role == 2:
        team_leader = TeamLeader.objects.get(auth_user=request.user)
        agents = Agent.objects.filter(team_leader=team_leader)
    else:
        agents = Agent.objects.all()

    for row_num, agent in enumerate(agents, 2):
        # Get the related data for each agent
        leads_count = Lead.objects.filter(assigned_to_agents=agent).count()
        paid_customers_count = PaidCustomer.objects.filter(lead__assigned_to_agents=agent).count()
        tasks_completed = Task.objects.filter(assigned_to=agent.auth_user,status='completed').count()
       
        resolved_complaint = Complaint.objects.filter(assigned_to=agent.auth_user,resolved=True).count()
        leaves_taken_count = AgentAttendance.objects.filter(agent=agent,status='Absent').count()

        # Format the joining date as a string in the desired format (e.g., "YYYY-MM-DD")
        joining_date_str = agent.joining_date.strftime("%Y-%m-%d")

        # Write the data for each agent to the Excel
        data = [joining_date_str,
                f"{agent.first_name} {agent.last_name}",
                f"{agent.team_leader.first_name} {agent.team_leader.last_name}",
                leads_count,
                paid_customers_count,
                tasks_completed,
               
                resolved_complaint,
                leaves_taken_count]

        for col_num, value in enumerate(data, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(response)
    return response


@login_required
def teamleader_reports(request):

    teamleader_search_query = request.GET.get("teamleader_search", "")
    teamleader_date_from_str = request.GET.get("teamleader_date_from", "")
    teamleader_date_to_str = request.GET.get("teamleader_date_to", "")

    teamleaders = TeamLeader.objects.all()

    if teamleader_search_query:
        teamleaders = teamleaders.filter(
            Q(first_name__icontains=teamleader_search_query) |
            Q(last_name__icontains=teamleader_search_query) |
            Q(email__icontains=teamleader_search_query) |
            Q(phone_number__icontains=teamleader_search_query)
        )

    teamleader_data = []
    for teamleader in teamleaders:
        team_agents = Agent.objects.filter(team_leader=teamleader)
        leads = Lead.objects.filter(status='converted',assigned_to=teamleader)
        paid_customers = PaidCustomer.objects.filter(lead__assigned_to=teamleader,payment_status='completed')
        leaves_taken = TeamLeaderAttendance.objects.filter(team_leader=teamleader,status='Absent')
        working_days = TeamLeaderAttendance.objects.filter(
            Q(team_leader=teamleader, status='Present') | Q(team_leader=teamleader, status='Late')
        )
        tasks = Task.objects.filter(assigned_to=teamleader.auth_user,status='completed')
        complaints = Complaint.objects.filter(assigned_to=teamleader.auth_user,resolved=True)

        if teamleader_date_from_str and teamleader_date_to_str:
            teamleader_date_from = datetime.strptime(teamleader_date_from_str, "%Y-%m-%d").date()
            teamleader_date_to = datetime.strptime(teamleader_date_to_str, "%Y-%m-%d").date()
            if teamleader.joining_date >= teamleader_date_from and teamleader.joining_date <= teamleader_date_to:
                teamleader_data.append({
                    'teamleader': teamleader,
                    'leads': leads,
                    'paid_customers': paid_customers,
                    'working_days':working_days,
                    'leaves_taken': leaves_taken,
                    'tasks':tasks,
                    'complaints':complaints,
                    'team_agents':team_agents
                })
        else:
            teamleader_data.append({
                'teamleader': teamleader,
                'leads': leads,
                'paid_customers': paid_customers,
                'working_days':working_days,
                'leaves_taken': leaves_taken,
                'tasks':tasks,
                'complaints':complaints,
                'team_agents':team_agents
            })

    context = {
        'teamleader_data': teamleader_data,
        'teamleader_search_query': teamleader_search_query,
        'teamleader_date_from': teamleader_date_from_str,  
        'teamleader_date_to': teamleader_date_to_str,      
    }

    return render(request, 'teamleader_reports.html', context)


@login_required
def teamleader_performance_report(request, teamleader_id):
    teamleader = get_object_or_404(TeamLeader, id=teamleader_id)
    team_agents = Agent.objects.filter(team_leader=teamleader)
    leads = Lead.objects.filter(assigned_to=teamleader)
    paid_customers = PaidCustomer.objects.filter(lead__assigned_to=teamleader)
    leaves_taken = TeamLeaderAttendance.objects.filter(team_leader=teamleader,status='Absent')
    tasks = Task.objects.filter(assigned_to=teamleader.auth_user)
    complaints = Complaint.objects.filter(assigned_to=teamleader.auth_user)

    context = {
        'teamleader': teamleader,
        'leads': leads,
        'paid_customers': paid_customers,
        'leaves_taken': leaves_taken,
        'tasks':tasks,
        'complaints':complaints,
        'team_agents':team_agents
    }

    return render(request, 'teamleader_performance_report.html', context)


@login_required
def teamleader_attendance_report(request, teamleader_id):
    teamleader = get_object_or_404(TeamLeader, id=teamleader_id)
    attendance = TeamLeaderAttendance.objects.filter(team_leader=teamleader)
    present_attendance = TeamLeaderAttendance.objects.filter(team_leader=teamleader, status='Present')
    absent_attendance = TeamLeaderAttendance.objects.filter(team_leader=teamleader, status='Absent')
    late_attendance = TeamLeaderAttendance.objects.filter(team_leader=teamleader, status='Late')
    total_working_days = present_attendance.count() + late_attendance.count()

    teamleader_date_from_str = request.GET.get("teamleader_date_from", "")
    teamleader_date_to_str = request.GET.get("teamleader_date_to", "")

    if teamleader_date_from_str and teamleader_date_to_str:
        teamleader_date_from = timezone.datetime.strptime(teamleader_date_from_str, "%Y-%m-%d")
        teamleader_date_to = timezone.datetime.strptime(teamleader_date_to_str, "%Y-%m-%d")
        attendance = attendance.filter(date__range=(teamleader_date_from, teamleader_date_to))
        present_attendance = present_attendance.filter(date__range=(teamleader_date_from, teamleader_date_to))
        absent_attendance = absent_attendance.filter(date__range=(teamleader_date_from, teamleader_date_to))
        late_attendance = late_attendance.filter(date__range=(teamleader_date_from, teamleader_date_to))
        total_working_days = present_attendance.count() + late_attendance.count()

    sort_by = request.GET.get("sort")
    if sort_by == "date":
        attendance = attendance.order_by("date")
    elif sort_by == "status":
        attendance = attendance.order_by("status")

    filter_by = request.GET.get("filter")
    if filter_by == "today":
        today = timezone.now().date()
        attendance = attendance.filter(date=today)
    elif filter_by == "this_week":
        start_of_week = timezone.now().date() - timezone.timedelta(days=timezone.now().weekday())
        end_of_week = start_of_week + timezone.timedelta(days=6)
        attendance = attendance.filter(date__range=[start_of_week, end_of_week])
    elif filter_by == "this_month":
        today = timezone.now().date()
        start_of_month = today.replace(day=1)
        end_of_month = start_of_month.replace(day=calendar.monthrange(start_of_month.year, start_of_month.month)[1])
        attendance = attendance.filter(date__range=[start_of_month, end_of_month])
    elif filter_by == "6_months":
        today = timezone.now().date()
        start_of_six_months_ago = today - timezone.timedelta(days=180)
        attendance = attendance.filter(date__gte=start_of_six_months_ago)
    elif filter_by == "this_year":
        today = timezone.now().date()
        start_of_year = today.replace(month=1, day=1)
        end_of_year = today.replace(month=12, day=31)
        attendance = attendance.filter(date__range=[start_of_year, end_of_year])

    context = {
        'teamleader': teamleader,
        'attendance': attendance,
        'present_attendance': present_attendance,
        'absent_attendance': absent_attendance,
        'late_attendance': late_attendance,
        'total_working_days': total_working_days,
        'teamleader_date_from': teamleader_date_from_str,
        'teamleader_date_to': teamleader_date_to_str,
    }

    return render(request, 'teamleader_attendance_report.html', context)


@login_required
def export_teamleader_attendance(request, teamleader_id):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filter_value = request.GET.get("filter", "")
    
    # Fetch the teamleader data
    teamleader = get_object_or_404(TeamLeader, id=teamleader_id)

    # Set the Excel file name
    response['Content-Disposition'] = f'attachment; filename="{teamleader.first_name}_{teamleader.last_name}_teamleader_AttendanceReport.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the headers for the first table (teamleader Information)
    headers_teamleader = ['Teamleader Name', 'Phone Number', 'Total Working Days', 'Days Present', 'Days Late', 'Days Absent']
    for col_num, header in enumerate(headers_teamleader, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Fetch the attendance data for the specific teamleader
    teamleader_attendance = TeamLeaderAttendance.objects.filter(team_leader=teamleader)
    present_attendance = teamleader_attendance.filter(status='Present')
    absent_attendance = teamleader_attendance.filter(status='Absent')
    late_attendance = teamleader_attendance.filter(status='Late')
    total_working_days = present_attendance.count() + late_attendance.count()

    # Write the data for the first table (teamleader Information)
    data_teamleader = [f"{teamleader.first_name} {teamleader.last_name}",
                  teamleader.phone_number,
                  total_working_days,
                  present_attendance.count(),
                  late_attendance.count(),
                  absent_attendance.count()]

    for col_num, value in enumerate(data_teamleader, 1):
        worksheet.cell(row=2, column=col_num, value=value)

    # Write the headers for the second table (Attendance Data)
    headers_attendance = ['Date', 'Status', 'Working Place', 'Login Time', 'Logout Time']
    for col_num, header in enumerate(headers_attendance, 1):
        worksheet.cell(row=4, column=col_num, value=header)

    # Fetch the attendance data for the specific teamleader based on filters (if any)
    if filter_value == 'today':
        attendance_data = TeamLeaderAttendance.objects.filter(team_leader=teamleader, date=date.today())
    elif filter_value == 'this_week':
        start_date = date.today() - timedelta(days=date.today().weekday())
        end_date = start_date + timedelta(days=6)
        attendance_data = TeamLeaderAttendance.objects.filter(team_leader=teamleader, date__range=[start_date, end_date])
    elif filter_value == 'this_month':
        start_date = date.today().replace(day=1)
        end_date = date.today().replace(day=calendar.monthrange(date.today().year, date.today().month)[1])
        attendance_data = TeamLeaderAttendance.objects.filter(team_leader=teamleader, date__range=[start_date, end_date])
    elif filter_value == '6_months':
        today = date.today()
        start_date = today - timedelta(days=180)
        end_date = today
        attendance_data = TeamLeaderAttendance.objects.filter(team_leader=teamleader, date__range=[start_date, end_date])
    elif filter_value == 'this_year':
        start_date = date.today().replace(month=1, day=1)
        end_date = date.today().replace(month=12, day=31)
        attendance_data = TeamLeaderAttendance.objects.filter(team_leader=teamleader, date__range=[start_date, end_date])
    else:
        attendance_data = TeamLeaderAttendance.objects.filter(team_leader=teamleader)

    for row_num, attendance in enumerate(attendance_data, 5):
        # Format the date as a string in the desired format (e.g., "YYYY-MM-DD")
        date_str = attendance.date.strftime("%Y-%m-%d")

        # Format login_time and logout_time only if they are not None
        login_time_str = attendance.login_time.strftime("%H:%M:%S") if attendance.login_time else ""
        logout_time_str = attendance.logout_time.strftime("%H:%M:%S") if attendance.logout_time else ""

        # Write the attendance data to the second table
        data_attendance = [date_str,
                           attendance.status,
                           attendance.working_place,
                           login_time_str,
                           logout_time_str,
                           ]

        for col_num, value in enumerate(data_attendance, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(response)
    return response


@login_required
def export_teamleader_reports(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="teamleader_report.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the headers for the Excel
    headers = ['Joining Date', 'Teamleader Name', 'Team Agents', 'Leads Converted', 'Paid Customers', 'Tasks Completed', 'Resolved Complaints', 'Leaves Taken']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Fetch all the teamleaders
    teamleaders = TeamLeader.objects.all()

    for row_num, teamleader in enumerate(teamleaders, 2):
        # Get the related data for each teamleader
        team_agents = Agent.objects.filter(team_leader=teamleader).count()
        leads_count = Lead.objects.filter(assigned_to=teamleader,status='converted').count()
        paid_customers_count = PaidCustomer.objects.filter(lead__assigned_to=teamleader,payment_status='completed').count()
        tasks_completed = Task.objects.filter(assigned_to=teamleader.auth_user,status='completed').count()
        resolved_complaint = Complaint.objects.filter(assigned_to=teamleader.auth_user,resolved=True).count()
        leaves_taken_count = TeamLeaderAttendance.objects.filter(team_leader=teamleader,status='Absent').count()

        team_agents = team_agents or 0
        leads_count = leads_count or 0
        paid_customers_count = paid_customers_count or 0
        tasks_completed = tasks_completed or 0
        resolved_complaint = resolved_complaint or 0
        leaves_taken_count = leaves_taken_count or 0


        # Format the joining date as a string in the desired format (e.g., "YYYY-MM-DD")
        joining_date_str = teamleader.joining_date.strftime("%Y-%m-%d")

        # Write the data for each teamleader to the Excel
        data = [joining_date_str,
                f"{teamleader.first_name} {teamleader.last_name}",
                team_agents,
                leads_count,
                paid_customers_count,
                tasks_completed,
                resolved_complaint,
                leaves_taken_count]

        for col_num, value in enumerate(data, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(response)
    return response

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import TeamLeaderAttendance, AgentAttendance
from django.contrib.auth.decorators import login_required

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import TeamLeaderAttendance, AgentAttendance
from django.contrib.auth.decorators import login_required
import json

@require_POST
@login_required
def update_break_status(request):
    user = request.user

    try:
        data = json.loads(request.body)
        is_on_break = data.get('is_on_break', False)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data.'})

    if user.role == 2:  # Assuming role is stored in user_profile
        try:
            # Update the TeamLeaderAttendance model
            team_leader_attendance = TeamLeaderAttendance.objects.get(team_leader=user.teamleader, is_on_break=is_on_break)
            team_leader_attendance.is_on_break = not is_on_break  # Toggle the value
            team_leader_attendance.save()
            return JsonResponse({'success': True, 'message': 'Team Leader break status updated.'})
        except TeamLeaderAttendance.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Team Leader attendance not found.'})
    elif user.role == 3:  # Assuming role is stored in user_profile
        try:
            # Update the AgentAttendance model
            agent_attendance = AgentAttendance.objects.get(agent=user.agent, is_on_break=is_on_break)
            agent_attendance.is_on_break = not is_on_break  # Toggle the value
            agent_attendance.save()
            return JsonResponse({'success': True, 'message': 'Agent break status updated.'})
        except AgentAttendance.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Agent attendance not found.'})
    else:
        return JsonResponse({'success': False, 'message': 'User role is not supported.'})







class TeamLeaderAttendanceListView(ListView):
    template_name = 'team_leader_attendance.html'
    model = TeamLeaderAttendance

    def get_queryset(self):
        team_leader_search_query = self.request.GET.get("team_leader_search", "")
        team_leader_date_from = self.request.GET.get("team_leader_date_from", "")
        team_leader_date_to = self.request.GET.get("team_leader_date_to", "")

        team_leaders = TeamLeaderAttendance.objects.all()

        # Format login and logout times for team leaders in local time zone
        for team_leader in team_leaders:
            team_leader.login_time = team_leader.login_time.strftime("%I:%M %p") if team_leader.login_time else "N/A"
            team_leader.logout_time = team_leader.logout_time.strftime("%I:%M %p") if team_leader.logout_time else "N/A"

        # Filter team leaders based on the search query (by first name and last name)
        if team_leader_search_query:
            team_leaders = team_leaders.filter(
                Q(team_leader__first_name__icontains=team_leader_search_query) |
                Q(team_leader__last_name__icontains=team_leader_search_query)
            )

        team_leaders = team_leaders.order_by('-date')

        # Filter team leaders based on the date range
        if team_leader_date_from and team_leader_date_to:
            team_leader_date_from = datetime.strptime(team_leader_date_from, "%Y-%m-%d")
            team_leader_date_to = datetime.strptime(team_leader_date_to, "%Y-%m-%d")
            team_leaders = team_leaders.filter(date__range=[team_leader_date_from, team_leader_date_to])

        filter_by = self.request.GET.get("filter")
        if filter_by == "today":
            today = timezone.now().date()
            team_leaders = team_leaders.filter(date=today)
        elif filter_by == "this_week":
            start_of_week = timezone.now().date() - timezone.timedelta(days=timezone.now().weekday())
            end_of_week = start_of_week + timezone.timedelta(days=6)
            team_leaders = team_leaders.filter(date__range=[start_of_week, end_of_week])
        elif filter_by == "this_month":
            today = timezone.now().date()
            start_of_month = today.replace(day=1)
            end_of_month = start_of_month.replace(day=calendar.monthrange(start_of_month.year, start_of_month.month)[1])
            team_leaders = team_leaders.filter(date__range=[start_of_month, end_of_month])
        elif filter_by == "6_months":
            today = timezone.now().date()
            start_of_six_months_ago = today - timezone.timedelta(days=180)
            team_leaders = team_leaders.filter(date__gte=start_of_six_months_ago)
        elif filter_by == "this_year":
            today = timezone.now().date()
            start_of_year = today.replace(month=1, day=1)
            end_of_year = today.replace(month=12, day=31)
            team_leaders = team_leaders.filter(date__range=[start_of_year, end_of_year])
       
        for team_leader in team_leaders:
            team_leader.break_objects = Break.objects.filter(username=team_leader.team_leader.username,start_time__date=team_leader.date)


        return team_leaders

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["team_leader_search_query"] = self.request.GET.get("team_leader_search", "")
        context["team_leader_date_from"] = self.request.GET.get("team_leader_date_from", "")
        context["team_leader_date_to"] = self.request.GET.get("team_leader_date_to", "")
        return context


class AgentAttendanceListView(ListView):
    template_name = 'agent_attendance.html'
    
    model = AgentAttendance

    def get_queryset(self):
        agent_search_query = self.request.GET.get("agent_search", "")
        agent_date_from = self.request.GET.get("agent_date_from", "")
        agent_date_to = self.request.GET.get("agent_date_to", "")

        current_user = self.request.user
        if current_user.is_authenticated and current_user.role == 2:
            team_leader = current_user.teamleader
            agents = AgentAttendance.objects.filter(agent__team_leader=team_leader)
        else:
            agents = AgentAttendance.objects.all()

        # Format login and logout times for agents in local time zone
        for agent in agents:
            agent.login_time = agent.login_time.strftime("%I:%M %p") if agent.login_time else "N/A"
            agent.logout_time = agent.logout_time.strftime("%I:%M %p") if agent.logout_time else "N/A"

        # Filter agents based on the search query (by first name and last name)
        if agent_search_query:
            agents = agents.filter(
                Q(agent__first_name__icontains=agent_search_query) |
                Q(agent__last_name__icontains=agent_search_query)
            )

        # Filter agents based on the date range
        if agent_date_from and agent_date_to:
            agent_date_from = datetime.strptime(agent_date_from, "%Y-%m-%d")
            agent_date_to = datetime.strptime(agent_date_to, "%Y-%m-%d")

        agents = agents.order_by('-date')

        filter_by = self.request.GET.get("filter")
        if filter_by == "today":
            today = timezone.now().date()
            agents = agents.filter(date=today)
        elif filter_by == "this_week":
            start_of_week = timezone.now().date() - timezone.timedelta(days=timezone.now().weekday())
            end_of_week = start_of_week + timezone.timedelta(days=6)
            agents = agents.filter(date__range=[start_of_week, end_of_week])
        elif filter_by == "this_month":
            today = timezone.now().date()
            start_of_month = today.replace(day=1)
            end_of_month = start_of_month.replace(day=calendar.monthrange(start_of_month.year, start_of_month.month)[1])
            agents = agents.filter(date__range=[start_of_month, end_of_month])
        elif filter_by == "6_months":
            today = timezone.now().date()
            start_of_six_months_ago = today - timezone.timedelta(days=180)
            agents = agents.filter(date__gte=start_of_six_months_ago)
        elif filter_by == "this_year":
            today = timezone.now().date()
            start_of_year = today.replace(month=1, day=1)
            end_of_year = today.replace(month=12, day=31)
            agents = agents.filter(date__range=[start_of_year, end_of_year])

                # Fetch Break objects for each agent on their login_date
        for agent in agents:
            agent.break_objects = Break.objects.filter(username=agent.agent.username, start_time__date=agent.date)

        return agents

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["agent_search_query"] = self.request.GET.get("agent_search", "")
        context["agent_date_from"] = self.request.GET.get("agent_date_from", "")
        context["agent_date_to"] = self.request.GET.get("agent_date_to", "")

        return context

    
def my_attendance(request):
    if request.user.role == 2:  # Team Leader
        team_leader = request.user.teamleader  # Assuming teamleader field exists on the user model

        current_team_leader_attendance = TeamLeaderAttendance.objects.filter(team_leader=team_leader)

        for attendance in current_team_leader_attendance:
            attendance.breaks = Break.objects.filter(
                username=attendance.team_leader.username,
                start_time__date=attendance.date,
            ).order_by('start_time')

        context = {
            'current_team_leader_attendance': current_team_leader_attendance,
        }
    elif request.user.role == 3:  
        agent = request.user.agent  # Assuming agent field exists on the user model

        current_agent_attendance = AgentAttendance.objects.filter(agent=agent)

        for attendance in current_agent_attendance:
            attendance.breaks = Break.objects.filter(
                username=attendance.agent.username,
                start_time__date=attendance.date,
            ).order_by('start_time')

        context = {
            'current_agent_attendance': current_agent_attendance,
        }
    else:
       
        context = {}

    return render(request, 'my_attendance.html', context)


@csrf_exempt
@login_required
def save_break(request):
    if request.method == "POST":
        # Get the form data from the request
        start_time_str = request.POST.get('start_time')
        stop_time_str = request.POST.get('stop_time')
        duration_str = request.POST.get('duration')


        start_time = parse_datetime(start_time_str)
        stop_time = parse_datetime(stop_time_str)

        duration_microseconds = float(duration_str)
        duration = timedelta(seconds=duration_microseconds)


        if request.user.role == 1:
            role = 'AD'
        elif request.user.role == 2:
            role = 'TL'
        elif request.user.role == 3:
            role = 'AG'
        else:
            role = 'Unknown'

        new_break = Break(
            username=request.user,
            role=role,
            start_time=start_time,
            stop_time=stop_time,
            duration=duration,
        )

        new_break.save()  

        return JsonResponse({'status': 'success'})
    
   
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



def assign_leads(request):
    from_agent_id = None
    to_agent_id = None
    from_team_leader_id = None
    to_team_leader_id = None

    if request.method == 'POST':
        if 'action' in request.POST:
            action = request.POST['action']
            if action == 'assign':
                if request.user.role == 1:
                    team_leader_id = request.POST.get('team_leader')
                    if team_leader_id:
                        team_leader = TeamLeader.objects.get(id=team_leader_id)
                        leads_to_assign = Lead.objects.filter(id__in=request.POST.getlist('leadCheckbox'))

                        # Find leads that are already assigned to a team leader
                        leads_already_assigned = leads_to_assign.filter(assigned_to__isnull=False)

                        # If there are any leads already assigned, handle their current assignment
                        if leads_already_assigned:
                            for lead in leads_already_assigned:
                                # Find agents assigned to this lead
                                agents_assigned_to_lead = lead.assigned_to_agents
                                
                                # Remove their current assignment
                                agents_assigned_to_lead = None
                                lead.save()

                                # Create a TeamLeaderNotification for the lead assignment
                                TeamLeaderNotification.objects.create(
                                    user=team_leader.auth_user,
                                    team_leader=team_leader,
                                    message=f'Lead : {lead.name} has been assigned to {team_leader.first_name} {team_leader.last_name}.'
                                )

                        # Assign leads to the new team leader and set assigned_to_agents to None
                        leads_to_assign.update(assigned_to=team_leader, assigned_to_agents=None)
                        messages.success(request, 'Leads successfully assigned to the team leader.')
                    else:
                        messages.error(request, 'Please select a team leader.')

                elif request.user.role == 2:
                    agent_id = request.POST.get('agent')
                    if agent_id:
                        agent = Agent.objects.get(id=agent_id)
                        leads = Lead.objects.filter(id__in=request.POST.getlist('leadCheckbox'))
                        leads.update(assigned_to_agents=agent)

                        # Create a TeamLeaderNotification for each lead assignment
                        for lead in leads:
                            TeamLeaderNotification.objects.create(
                                user=agent.auth_user,
                                agent=agent,
                                message=f'Lead "{lead.name}" has been assigned to {agent.first_name} {agent.last_name}.'
                            )

                        messages.success(request, 'Leads successfully assigned to the agent.')
                    else:
                        messages.error(request, 'Please select an agent.')

            elif action == 'transfer':
                # Transfer leads between agents or team leaders
                if request.user.role == 3:
                    from_agent_id = request.user.agent.id
                    to_agent_id = request.POST.get('agent')
                elif request.user.role == 2:
                    from_team_leader_id = request.user.teamleader.id
                    to_team_leader_id = request.POST.get('team_leader')

                leads = Lead.objects.filter(id__in=request.POST.getlist('leadCheckbox'))

                if from_agent_id and to_agent_id:
                    from_agent = Agent.objects.get(id=from_agent_id)
                    to_agent = Agent.objects.get(id=to_agent_id)
                    leads.update(assigned_to_agents=to_agent)

                    for lead in leads:
                        agent_lead_transfer = LeadTransfer.objects.create(
                            lead=lead,
                            from_agent=from_agent,
                            to_agent=to_agent,
                        )
            
                    messages.success(request, 'Leads successfully transferred between agents.')
                elif from_team_leader_id and to_team_leader_id:
                    from_team_leader = TeamLeader.objects.get(id=from_team_leader_id)
                    to_team_leader = TeamLeader.objects.get(id=to_team_leader_id)
                    for lead in leads:
                        if lead.assigned_to_agents:
                            lead.assigned_to_agents = None

                        team_leader_lead_transfer = LeadTransfer.objects.create(
                                lead=lead,
                                from_team_leader=from_team_leader,
                                to_team_leader=to_team_leader,
                            )

                        lead.assigned_to = to_team_leader
                        lead.save()
                    messages.success(request, 'Leads successfully transferred between team leaders.')
                else:
                    messages.error(request, 'Please select appropriate agents or team leaders for the transfer.')

        return redirect('lead_list')

    # Fetch required data for rendering the view
    agents = Agent.objects.filter(team_leader=request.user.teamleader) if request.user.role == 2 else Agent.objects.all()
    team_leaders = TeamLeader.objects.all()
    object_list = Lead.objects.all()

    context = {
        'agents': agents,
        'team_leaders': team_leaders,
        'object_list': object_list,
        'agent_lead_transfer':agent_lead_transfer,
        'team_leader_lead_transfer':team_leader_lead_transfer,
    }

    return render(request, 'assign_leads.html', context)


def myprofile(request):
    if request.user.is_authenticated:
        current_user = request.user
        if current_user.role == 2:  # Team Leader
            team_leader = TeamLeader.objects.get(auth_user=current_user)
            # Get related information for the team leader
            # Example:
            team_members = Agent.objects.filter(team_leader=team_leader)
            
            # Pass the necessary information to the template context
            context = {
                'team_leader': team_leader,
                'team_members': team_members,
            }
            return render(request, 'team_leader_profile.html', context)
        elif current_user.role == 3:  # Agent
            agent = Agent.objects.get(auth_user=current_user)
            # Get related information for the agent
            # Example:
            team_leader = agent.team_leader
            # Pass the necessary information to the template context
            context = {
                'agent': agent,
                'team_leader': team_leader,
            }
            return render(request, 'agent_profile.html', context)

    return HttpResponse("Unauthorized", status=401)



class TeamLeaderListView(LoginRequiredMixin, ListView):
    model = TeamLeader
    template_name = 'team_leader_list.html'
    paginate_by = 10

    def get_queryset(self):
        filter_val = self.request.GET.get("filter", "")
        order_by = self.request.GET.get("orderby", "-id")
        date_from = self.request.GET.get("date_from", "")
        date_to = self.request.GET.get("date_to", "")

        if filter_val:
            team_leaders = TeamLeader.objects.filter(
                Q(auth_user__username__icontains=filter_val) |
                Q(auth_user__email__icontains=filter_val) |
                Q(phone_number__icontains=filter_val)
            )
        else:
            team_leaders = TeamLeader.objects.all()

        if date_from and date_to:
            # Convert the date strings to datetime objects
            date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            date_to = datetime.strptime(date_to, "%Y-%m-%d").date()

            team_leaders = team_leaders.filter(
                joining_date__range=(date_from, date_to))

        team_leaders = team_leaders.order_by(order_by)

        return team_leaders

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")
        context["date_from"] = self.request.GET.get("date_from", "")
        context["date_to"] = self.request.GET.get("date_to", "")
        context["all_table_fields"] = TeamLeader._meta.get_fields()
        return context


class TeamLeaderCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = TeamLeader
    fields = ['username', 'password', 'profile_pic', 'first_name', 'last_name',
              'email', 'phone_number', 'address', 'joining_date', 'remarks']
    template_name = 'team_leader_list.html'
    success_url = reverse_lazy('team_leader_list')
    success_message = "Team Leader Added!"

    def form_valid(self, form):
        User = get_user_model()
        username = form.cleaned_data['username']
        
        # Check if the username already exists in the CustomUser table
        if CustomUser.objects.filter(username=username).exists():
            messages.error(self.request, "Username already exists.")
            return self.form_invalid(form)

        user = User.objects.create_user(
            username=username,
            password=form.cleaned_data['password']
        )
        user.role = 2  # Set role as 2 for team leader
        user.save()
        self.object = form.save(commit=False)
        self.object.auth_user = user
        self.object.save()
        return super().form_valid(form)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['profile_pic'].required = False
        return form

    def form_invalid(self, form):
        for field in form.errors:
            error_messages = form.errors[field]
            for error_message in error_messages:
                messages.error(self.request, f"{field.capitalize()} : {error_message}")
        return HttpResponseRedirect(reverse_lazy('team_leader_list'))


class TeamLeaderUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = TeamLeader
    fields = ['username', 'password', 'profile_pic', 'first_name', 'last_name',
              'email', 'phone_number', 'address', 'joining_date', 'remarks']
    template_name = 'team_leader_update.html'
    success_url = reverse_lazy('team_leader_list')
    success_message = "Team Leader Updated!"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if self.request.user.role == 2:
            form.fields = {'username': forms.CharField(), 'password': forms.CharField(),'profile_pic':forms.ImageField(),
                           'first_name':forms.CharField(),'last_name':forms.CharField(),'email':forms.EmailField(),
                           'phone_number':forms.CharField()}
        return form
    
    def form_valid(self, form):
        if self.request.user.role == 2 and self.request.user.teamleader == self.object:
            self.success_url = reverse_lazy('team_leader_update', kwargs={'pk': self.object.pk})
        return super().form_valid(form)


class TeamLeaderDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = TeamLeader
    template_name = 'team_leader_list.html'
    success_url = reverse_lazy('team_leader_list')
    success_message = "Team Leader Deleted!"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        team_leader = self.object

        custom_user = team_leader.auth_user  # Get the associated CustomUser
        self.object.delete()  # Delete the team leader

        if custom_user:
            custom_user.delete()  # Delete the associated CustomUser

        return redirect(self.get_success_url())

    def get_object(self, queryset=None):
        obj = super().get_object(queryset=queryset)
        return obj.auth_user  # Return t

class AgentListView(LoginRequiredMixin, ListView):
    model = Agent
    template_name = 'agent_list.html'
    paginate_by = 10

    def get_queryset(self):
        filter_val = self.request.GET.get("filter", "")
        date_from = self.request.GET.get("date_from", "")
        date_to = self.request.GET.get("date_to", "")
        team_leader = self.request.GET.get("team_leader", "")
        order_by = self.request.GET.get("orderby", "-id")

        agents = Agent.objects.all()

        user = self.request.user

        if user.role == 2:  # Assuming role field exists on the user model
            team_leader = user.teamleader  # Assuming teamleader field exists on the user model

            agents = Agent.objects.filter(team_leader=team_leader)


        if filter_val:
            agents = agents.filter(
                Q(auth_user__username__icontains=filter_val) |
                Q(id__icontains=filter_val) |
                Q(email__icontains=filter_val) |
                Q(phone_number__icontains=filter_val)
            )

        if date_from and date_to:
            # Convert the date strings to datetime objects
            date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            date_to = datetime.strptime(date_to, "%Y-%m-%d").date()

            agents = agents.filter(joining_date__range=(date_from, date_to))

        if team_leader:
            agents = agents.filter(team_leader_id=team_leader)

        agents = agents.order_by(order_by)

        return agents

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["date_from"] = self.request.GET.get("date_from", "")
        context["date_to"] = self.request.GET.get("date_to", "")
        context["team_leader"] = self.request.GET.get("team_leader", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")
        context["team_leaders"] = TeamLeader.objects.all()
        return context



class AgentCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Agent
    fields = ['username', 'password', 'profile_pic', 'first_name', 'last_name',
              'email', 'phone_number', 'address', 'joining_date', 'remarks', 'team_leader']
    template_name = 'agent_list.html'
    success_url = reverse_lazy('agent_list')
    success_message = "Agent Added!"

    def form_valid(self, form):
        User = get_user_model()
        user = User.objects.create_user(
            username=form.cleaned_data['username'],
            password=form.cleaned_data['password']
        )
        user.role = 3  
        user.save()
        self.object = form.save(commit=False)
        self.object.auth_user = user
        self.object.save()
        return super().form_valid(form)
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['profile_pic'].required = False
        return form
    
    def form_invalid(self, form):
        for field in form.errors:
            error_messages = form.errors[field]
            for error_message in error_messages:
                messages.error(self.request, f"{field.capitalize()} : {error_message}")
        return HttpResponseRedirect(reverse_lazy('agent_list'))
    
        
        
class AgentUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Agent
    fields = ['username', 'password', 'profile_pic', 'first_name', 'last_name',
              'email', 'phone_number', 'address', 'joining_date', 'remarks', 'team_leader']
    template_name = 'agent_update.html'
    success_url = reverse_lazy('agent_list')
    success_message = "Agent Updated!"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if self.request.user.role == 3:
            form.fields = {'username': forms.CharField(), 'password': forms.CharField(),'profile_pic':forms.ImageField(),
                           'first_name':forms.CharField(),'last_name':forms.CharField(),'email':forms.EmailField(),
                           'phone_number':forms.CharField()}
        return form
    
    def get_success_url(self):
        if self.request.user.role == 3:
            return reverse('agent_update', kwargs={'pk': self.request.user.agent.id})
        return super().get_success_url()


class AgentDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Agent
    template_name = 'agent_list.html'
    success_url = reverse_lazy('agent_list')
    success_message = "Agent Deleted!"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        agent = self.object

        custom_user = agent.auth_user  # Get the associated CustomUser
        self.object.delete()  # Delete the team leader

        if custom_user:
            custom_user.delete()  # Delete the associated CustomUser

        return redirect(self.get_success_url())

    def get_object(self, queryset=None):
        obj = super().get_object(queryset=queryset)
        return obj.auth_user  # Return 


class LeadTransferListView(ListView):
    model = LeadTransfer
    template_name = 'lead_transfer_list.html'  
    context_object_name = 'lead_transfers'

    def get_queryset(self):
        queryset = LeadTransfer.objects.all()

        # Filter out transfers where from_agent and to_agent, or from_team_leader and to_team_leader are the same
        queryset = queryset.filter(
            ~Q(from_agent=F('to_agent')) | Q(to_agent__isnull=True),
            ~Q(from_team_leader=F('to_team_leader')) | Q(to_team_leader__isnull=True)
        )

        for lead_transfer in queryset:
            transfer_date = lead_transfer.transfer_date
            lead_transfer.local_date = transfer_date.date()
            lead_transfer.local_time = transfer_date.time()

        return queryset

    def post(self, request, *args, **kwargs):
        if 'delete_transfer' in request.POST:
            transfer_id = request.POST.get('delete_transfer')
            try:
                transfer = LeadTransfer.objects.get(pk=transfer_id)
                transfer.delete()
            except LeadTransfer.DoesNotExist:
                pass

        return redirect('lead_transfer_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")

        if self.request.user.is_authenticated:
            current_user = self.request.user
            LeadTransferNotification.objects.filter(user=current_user).update(seen=True)

        return context


def export_lead_report(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Fetch the selected filters from the request
    filter_date_from = request.GET.get("date_from", "")
    filter_date_to = request.GET.get("date_to", "")
    filter_source = request.GET.get("source", "")
    filter_status = request.GET.get("status", "")
    filter_assigned_to = request.GET.get("assigned_to", "")
    filter_agent = request.GET.get("agent", "")
    
    # Set the Excel file name
    response['Content-Disposition'] = 'attachment; filename="lead_report.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the headers for the Excel
    headers = ['Date', 'Lead Name', 'Contact Number', 'Status', 'Disposition', 'Assigned To Team Leader', 'Assigned To Agent']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Fetch the leads data based on the selected filters
    leads = Lead.objects.all()

    if filter_date_from and filter_date_to:
        leads = leads.filter(date__range=(filter_date_from, filter_date_to))

    if filter_source:
        leads = leads.filter(source=filter_source)

    if filter_status:
        leads = leads.filter(status=filter_status)

    if filter_assigned_to:
        leads = leads.filter(assigned_to=filter_assigned_to)

    if filter_agent:
        leads = leads.filter(assigned_to_agents=filter_agent)

    for row_num, lead in enumerate(leads, 2):
        # Format the date as a string in the desired format (e.g., "YYYY-MM-DD")
        date_str = lead.date.strftime("%Y-%m-%d")

        follow_up_actions = str(lead.follow_up_actions)

        # Write the lead data to the Excel
        data = [date_str,
                lead.name,
                lead.contact_info.phone_number,
                lead.get_status_display(),
                follow_up_actions,
                f"{lead.assigned_to.first_name} {lead.assigned_to.last_name}" if lead.assigned_to else "",
                f"{lead.assigned_to_agents.first_name} {lead.assigned_to_agents.last_name}" if lead.assigned_to_agents else ""
                ]

        for col_num, value in enumerate(data, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(response)
    return response

from django.contrib.auth.decorators import login_required  
from .models import Lead, LeadReminder  

def lead_reminder(request, pk): 
    lead = get_object_or_404(Lead, pk=pk)
    current_user = request.user

    if request.method == 'POST':
        reminder_date = request.POST.get('reminder_date')
        notes = request.POST.get('notes')
        priority = request.POST.get('priority')

        LeadReminder.objects.create(
            lead=lead,
            set_by = current_user,
            reminder_date=reminder_date,
            notes=notes,
            priority=priority
        )
        return redirect("lead_list")

    reminders = LeadReminder.objects.filter(lead=lead) 

    context = {
        'lead': lead,  
        'reminders': reminders, 
    }

    return render(request,'lead_reminder.html',context)


from django.shortcuts import get_object_or_404
from django.http import JsonResponse

def clear_lead_reminders(request, lead_id):
    lead = get_object_or_404(Lead, pk=lead_id)
    
    # Clear reminders for the current lead
    lead.reminders.all().delete()
    
    return redirect('lead_reminder', pk=lead_id)



from django.db.models import Prefetch

class LeadListView(LoginRequiredMixin, ListView):
    model = Lead
    template_name = 'lead_list.html'
    paginate_by = 10

    def get_queryset(self):
        filter_val = self.request.GET.get("filter", "")
        status = self.request.GET.get("status", "")
        date_from = self.request.GET.get("date_from", "")
        date_to = self.request.GET.get("date_to", "")
        order_by = self.request.GET.get("orderby", "-id")
        date_filter = self.request.GET.get("date_filter", "")  

        leads = Lead.objects.all()

        if filter_val:
            leads = leads.filter(
                Q(name__icontains=filter_val) |
                Q(contact_info__phone_number__icontains=filter_val) |
                Q(contact_info__name__icontains=filter_val) |
                Q(contact_info__email__icontains=filter_val) |
                Q(assigned_to__first_name__icontains=filter_val) |
                Q(assigned_to__last_name__icontains=filter_val)
            )

        if status:
            leads = leads.filter(status=status)

        if date_from and date_to:
            leads = leads.filter(date__range=(date_from, date_to))

        # Filter leads assigned to the current team leader (role 2)
        if self.request.user.is_authenticated and self.request.user.role == 2:
            team_leader_id = self.request.user.teamleader.id
            agents = Agent.objects.filter(team_leader=team_leader_id)
            leads = leads.filter(Q(assigned_to=team_leader_id) | Q(assigned_to_agents__in=agents))

        # Filter leads assigned to the current agent (role 3)
        elif self.request.user.is_authenticated and self.request.user.role == 3:
            agent = Agent.objects.get(auth_user=self.request.user)
            leads = leads.filter(assigned_to_agents=agent)


        if date_filter == "today":
            today = date.today()
            leads = leads.filter(date=today)

        elif date_filter == "this_week":
            today = date.today()
            start_of_week = today - timedelta(days=today.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            leads = leads.filter(date__range=[start_of_week, end_of_week])

        elif date_filter == "this_month":
            today = date.today()
            start_of_month = date(today.year, today.month, 1)
            end_of_month = date(today.year, today.month + 1, 1) - timedelta(days=1)
            leads = leads.filter(date__range=[start_of_month, end_of_month])

        elif date_filter == "last_6_months":
            today = date.today()
            start_of_last_6_months = date(today.year, today.month - 6, 1)
            end_of_last_6_months = date(today.year, today.month, 1) - timedelta(days=1)
            leads = leads.filter(date__range=[start_of_last_6_months, end_of_last_6_months])

        elif date_filter == "this_year":
            today = date.today()
            start_of_year = date(today.year, 1, 1)
            end_of_year = date(today.year, 12, 31)
            leads = leads.filter(date__range=[start_of_year, end_of_year])

        leads = leads.order_by(order_by)
        leads = leads.select_related('follow_up_actions')

        leads = leads.prefetch_related(
            Prefetch('reminders', queryset=LeadReminder.objects.all(), to_attr='lead_reminders')
        )

        return leads

    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["status"] = self.request.GET.get("status", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")
        
        # Get all team leaders
        if self.request.user.is_authenticated and self.request.user.role == 2:
            team_leader_id = self.request.user.teamleader.id
            leads_assigned_to_team_leader = Lead.objects.filter(assigned_to=team_leader_id)

            agents = Agent.objects.filter(team_leader=team_leader_id)
            leads_assigned_to_agents = Lead.objects.filter(assigned_to_agents__in=agents)

            # Combine both sets of leads and remove duplicates
            leads = (leads_assigned_to_team_leader | leads_assigned_to_agents).distinct()

            context["leads"] = leads

            current_user = self.request.user.teamleader  # Assuming the teamleader field exists on the user model
            team_leader = TeamLeader.objects.get(id=current_user.id)

            # Get all the agents under the current team leader
            agents = Agent.objects.filter(team_leader=team_leader)

            # Set the agents for the context
            context["agents"] = agents
        
        elif self.request.user.is_authenticated and self.request.user.role == 3:

            current_user_team_leader = self.request.user.agent.team_leader

            agents = Agent.objects.filter(team_leader=current_user_team_leader).exclude(auth_user=self.request.user)
        
            context["agents"] = agents
        else:
            # If the user is not a team leader, display all the leads
            context["leads"] = self.get_queryset()

        # Get all team leaders
        context["team_leaders"] = TeamLeader.objects.exclude(auth_user=self.request.user)
        context["contacts"] = Contact.objects.all()
        context["assigned_to"] = Lead.objects.all()
        context["custom_follow_up"] = CustomFollowUp.objects.all()
        users = CustomUser.objects.exclude(id=self.request.user.id)

        for user in users:
            try:
                team_leader = TeamLeader.objects.get(auth_user=user)
                user.team_leader_first_name = team_leader.first_name
                user.team_leader_last_name = team_leader.last_name
            except TeamLeader.DoesNotExist:
                user.team_leader_first_name = None
                user.team_leader_last_name = None

            try:
                agent = Agent.objects.get(auth_user=user)
                user.agent_first_name = agent.first_name
                user.agent_last_name = agent.last_name
            except Agent.DoesNotExist:
                user.agent_first_name = None
                user.agent_last_name = None

        context["users"] = users

        if self.request.user.is_authenticated:
            current_user = self.request.user
            LeadNotification.objects.filter(user=current_user).update(seen=True)

        return context
    

    def post(self, request, *args, **kwargs):
        if 'csv_file' in request.FILES:
            csv_file = request.FILES['csv_file']
            decoded_file = csv_file.read().decode('utf-8')
            csv_data = csv.reader(decoded_file.splitlines(), delimiter=',')
            next(csv_data)  # Skip the header row if present in the CSV

            for row in csv_data:
                date_str, name, contact_info_str, source, details = row
                date = datetime.strptime(date_str, "%d-%m-%Y").date()

                # Try to get an existing Contact instance by matching contact_info_str
                try:
                    contact_info = Contact.objects.get(phone_number=contact_info_str)
                except Contact.DoesNotExist:
                    # If the Contact doesn't exist, create a new one
                    contact_info = Contact.objects.create(phone_number=contact_info_str)

                # Create a Lead object and associate it with the Contact
                Lead.objects.create(
                    date=date,
                    name=name,
                    contact_info=contact_info,
                    source=source,
                    details=details
                )

            return redirect('lead_list')  # Assuming you have a URL named 'lead_list' for listing leads

        return self.get(request, *args, **kwargs)


def create_custom_follow_up(request):
    if request.method == 'POST' and request.user.is_authenticated and request.user.role == 1:
        follow_up_name = request.POST.get('follow_up_name')

        custom_follow_up = CustomFollowUp.objects.create(name=follow_up_name)

        return JsonResponse({
            'success': True,
            'follow_up_id': custom_follow_up.id,
        })
    else:
        return JsonResponse({'success': False})


class LeadCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Lead
    fields = ['date', 'name', 'contact_info', 'source', 'status', 'details', 'follow_up_actions']
    template_name = 'lead_list.html'
    success_url = reverse_lazy('lead_list')
    success_message = "Lead Created!"

    def form_valid(self, form):
        current_user = self.request.user

        if current_user.role == 2:  # Team Leader
            form.instance.assigned_to = current_user.teamleader
        elif current_user.role == 3:  # Agent
            form.instance.assigned_to = current_user.agent.team_leader
            form.instance.assigned_to_agents = current_user.agent

        # Save the lead instance
        self.object = form.save()

        # Now, create the follow-up history for this lead
        follow_up_date = self.object.date  # Assuming the follow-up date is the same as the lead's date
        follow_up_time = datetime.now().time()  # Using the current time as the follow-up time
        follow_up_actions = form.cleaned_data['follow_up_actions']
        notes = form.cleaned_data['details']  # You can use the 'details' field as the default notes for the follow-up

        FollowUpHistory.objects.create(
            lead=self.object,
            follow_up_date=follow_up_date,
            follow_up_time=follow_up_time,
            follow_up_actions=follow_up_actions,
            notes=notes,
        )

        return redirect(self.get_success_url())

    def form_invalid(self, form):
        for field in form.errors:
            error_message = form.errors[field]
            print(f"Field: {field}, Error: {error_message}")
        return super().form_invalid(form)
    
    def form_invalid(self, form):
        for field in form.errors:
            error_messages = form.errors[field]
            for error_message in error_messages:
                messages.error(self.request, f"{field.capitalize()} : {error_message}")
        return HttpResponseRedirect(reverse_lazy('lead_list'))
    

def LeadFollowUpHistoryView(request, pk):
    lead = get_object_or_404(Lead, id=pk)
    follow_up_history_entries = FollowUpHistory.objects.filter(lead=lead).order_by('-follow_up_date', '-follow_up_time')

    return render(request, 'follow_up_history_list.html', {'lead': lead, 'follow_up_history_entries': follow_up_history_entries})


class LeadUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Lead
    fields = ['date', 'name', 'contact_info', 'source',
              'status', 'details', 'follow_up_actions']
    template_name = 'lead_update.html'
    success_url = reverse_lazy('lead_list')
    success_message = "Lead Updated!"
    

    def form_valid(self, form):
        response = super().form_valid(form)

        # Check if the request is an AJAX request
        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            # Send JSON response for AJAX requests
            return JsonResponse({'message': self.success_message}, status=200)

        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Include the contacts in the context
        context["contacts"] = Contact.objects.all()
        return context

@method_decorator(csrf_exempt, name='dispatch')
class BulkLeadUpdateView(View):
    def post(self, request, *args, **kwargs):
        updated_leads_data = json.loads(request.body)

        for lead_data in updated_leads_data:
            lead_id = lead_data.get('id')
            lead_status = lead_data.get('status')
            lead_follow_up_actions_id = lead_data.get('follow_up_actions')  # Assuming it's the ID of CustomFollowUp
            lead_notes = lead_data.get('notes')

            # Get the lead from the database and update the fields
            lead = get_object_or_404(Lead, pk=lead_id)

            # Check if the lead's status, follow-up actions, or notes have changed
            if lead.status != lead_status or lead.follow_up_actions_id != lead_follow_up_actions_id or lead.details != lead_notes:
                # Create a new follow-up history for the lead
                follow_up_date = datetime.now() # Assuming the follow-up date is the same as the lead's date
                follow_up_time = datetime.now().time()  # Using the current time as the follow-up time

                # Determine the agent, team_leader, and admin for the new history based on the current user role
                admin = None
                team_leader = None
                agent = None

                if request.user.role == 1:
                    admin = request.user
                elif request.user.role == 2:
                    team_leader = request.user.teamleader
                elif request.user.role == 3:
                    agent = request.user.agent

                # Determine the follow-up actions for the new history
                try:
                    follow_up_actions = CustomFollowUp.objects.get(pk=lead_follow_up_actions_id)
                except CustomFollowUp.DoesNotExist:
                    follow_up_actions = None  # Handle the case where follow-up actions don't exist

                # Create the new follow-up history entry
                FollowUpHistory.objects.create(
                    lead=lead,
                    follow_up_date=follow_up_date,
                    follow_up_time=follow_up_time,
                    follow_up_actions=follow_up_actions,
                    notes=lead_notes,
                    agent=agent,
                    team_leader=team_leader,
                    admin=admin,
                )

            lead.status = lead_status
            lead.follow_up_actions_id = lead_follow_up_actions_id
            lead.details = lead_notes
            lead.save()

        return JsonResponse({'message': 'Bulk leads updated successfully'}, status=200)


def delete_lead_view(request, lead_id):
    if request.method == 'DELETE':
        lead = get_object_or_404(Lead, pk=lead_id)
        lead.delete()
        return JsonResponse({'message': 'Lead deleted successfully.'})
    else:
        return HttpResponse(status=405)  # Method Not Allowed for other HTTP methods


class ContactListView(LoginRequiredMixin, ListView):
    model = Contact
    template_name = 'contact_list.html'
    paginate_by = 50
    def get_queryset(self):
        filter_val = self.request.GET.get("filter", "")
        date_from = self.request.GET.get("date_from", "")
        date_to = self.request.GET.get("date_to", "")
        order_by = self.request.GET.get("orderby", "-id")

        contacts = Contact.objects.all()

        if filter_val:
            contacts = contacts.filter(
                Q(name__icontains=filter_val) |
                Q(id__icontains=filter_val) |
                Q(date__icontains=filter_val)
            )


        if date_from and date_to:
            # Convert the date strings to datetime objects
            date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            date_to = datetime.strptime(date_to, "%Y-%m-%d").date()

            contacts = contacts.filter(date__range=(date_from, date_to))

        contacts = contacts.order_by(order_by)

        return contacts

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")

        if self.request.user.is_authenticated:
            current_user = self.request.user
            ContactNotification.objects.filter(user=current_user).update(seen=True)

        return context


class ContactCreateView(LoginRequiredMixin, CreateView):
    model = Contact
    fields = ['date', 'name', 'email', 'phone_number',
              'address', 'occupation', 'notes']
    template_name = 'contact_list.html'
    success_url = reverse_lazy('contact_list')
    success_message = "Contact Created!"

    def form_valid(self, form):
        self.object = form.save()
        return redirect(self.get_success_url())
    
    def form_invalid(self, form):
        for field in form.errors:
            error_messages = form.errors[field]
            for error_message in error_messages:
                messages.error(self.request, f"{field.capitalize()} : {error_message}")
        return HttpResponseRedirect(reverse_lazy('contact_list'))


class ContactUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Contact
    fields = ['date', 'name', 'email', 'phone_number',
              'address', 'occupation', 'notes']
    template_name = 'contact_update.html'
    success_url = reverse_lazy('contact_list')
    success_message = "Contact Updated!"


@method_decorator(csrf_exempt, name='dispatch')
class UpdateContactNotesView(View):
    def post(self, request, *args, **kwargs):
        updated_contacts_data = json.loads(request.body)
        for contact_data in updated_contacts_data:
            contact_id = contact_data.get('id')
            contact_notes = contact_data.get('notes')

            # Get the contact from the database and update the notes field
            try:
                contact = Contact.objects.get(pk=contact_id)
                contact.notes = contact_notes
                contact.save()
            except Contact.DoesNotExist:
                pass  # Handle the case where the contact doesn't exist (optional)

        return JsonResponse({'message': 'Contact notes updated successfully'}, status=200)



class ContactDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Contact
    template_name = 'contact_list.html'
    success_url = reverse_lazy('contact_list')
    success_message = "Contact Deleted!"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        return redirect(self.get_success_url())


def update_verify_request(request, paidcustomer_id):
    if request.method == 'GET':
        # Retrieve the PaidCustomer object
        paidcustomer = get_object_or_404(PaidCustomer, pk=paidcustomer_id)

        # Update the verify_request field to False
        paidcustomer.verify_request = True
        paidcustomer.save()

        # Redirect back to the previous page or a specific URL
        return redirect('paidcustomer_list')  # Redirect to the list of PaidCustomers or adjust as needed

def update_verifed(request, paidcustomer_id):
    if request.method == 'GET':
        # Retrieve the PaidCustomer object
        paidcustomer = get_object_or_404(PaidCustomer, pk=paidcustomer_id)

        # Update the verify_request field to False
        paidcustomer.verified = True
        paidcustomer.save()

        # Redirect back to the previous page or a specific URL
        return redirect('paidcustomer_list')  # Redirect to the list of PaidCustomers or adjust as needed



class PaidCustomerListView(LoginRequiredMixin, ListView):
    model = PaidCustomer
    template_name = 'paidcustomer_list.html'
    paginate_by = 10

    def get_queryset(self):
        filter_val = self.request.GET.get("filter", "")
        payment_status = self.request.GET.get("payment_status", "")
        date_from = self.request.GET.get("date_from", "")
        date_to = self.request.GET.get("date_to", "")
        order_by = self.request.GET.get("orderby", "-id")

        # Start with all PaidCustomer instances
        paid_customers = PaidCustomer.objects.all()

        # Apply general filters
        if filter_val:
            paid_customers = paid_customers.filter(
                Q(related_contact__name__icontains=filter_val) |
                Q(related_contact__phone_number__icontains=filter_val) |
                Q(lead__assigned_to_agents__first_name__icontains=filter_val) |
                Q(lead__assigned_to_agents__last_name__icontains=filter_val) |
                Q(payment_method__icontains=filter_val)
            )

        if payment_status:
            paid_customers = paid_customers.filter(payment_status=payment_status)

        if date_from and date_to:
            paid_customers = paid_customers.filter(payment_date__range=(date_from, date_to))

        # Check user roles and filter accordingly
        if self.request.user.is_authenticated and self.request.user.role == 2:
            team_leader = TeamLeader.objects.get(auth_user=self.request.user)
            agent_ids = Agent.objects.filter(team_leader=team_leader).values_list('auth_user', flat=True)
            filtered_paid_customers = PaidCustomer.objects.filter(
                Q(lead__assigned_to=self.request.user.teamleader.id) | Q(lead__assigned_to_agents__auth_user__in=agent_ids)
            )
        elif self.request.user.role == 3:
            agent = Agent.objects.get(auth_user=self.request.user)
            team_leader = agent.team_leader
            filtered_paid_customers = PaidCustomer.objects.filter(
                lead__assigned_to_agents=self.request.user.agent
            )
        else:
            filtered_paid_customers = paid_customers

        # Apply search filter separately on all PaidCustomer instances
        if filter_val:
            search_query = Q(
                related_contact__name__icontains=filter_val) | \
                Q(related_contact__phone_number__icontains=filter_val) | \
                Q(lead__assigned_to_agents__first_name__icontains=filter_val) | \
                Q(lead__assigned_to_agents__last_name__icontains=filter_val) | \
                Q(payment_method__icontains=filter_val)

            paid_customers = paid_customers.filter(search_query)
        else:
            # If no search filter, use filtered_paid_customers
            paid_customers = filtered_paid_customers
        


        filter_by = self.request.GET.get("filter_by")
        if filter_by == "today":
            today = timezone.now().date()
            paid_customers = paid_customers.filter(payment_date=today)
        elif filter_by == "this_week":
            start_of_week = timezone.now().date() - timezone.timedelta(days=timezone.now().weekday())
            end_of_week = start_of_week + timezone.timedelta(days=6)
            paid_customers = paid_customers.filter(payment_date__range=[start_of_week, end_of_week])
        elif filter_by == "this_month":
            today = timezone.now().date()
            start_of_month = today.replace(day=1)
            end_of_month = start_of_month.replace(day=calendar.monthrange(start_of_month.year, start_of_month.month)[1])
            paid_customers = paid_customers.filter(payment_date__range=[start_of_month, end_of_month])
        elif filter_by == "6_months":
            today = timezone.now().date()
            start_of_six_months_ago = today - timezone.timedelta(days=180)
            paid_customers = paid_customers.filter(payment_date__gte=start_of_six_months_ago)
        elif filter_by == "this_year":
            today = timezone.now().date()
            start_of_year = today.replace(month=1, day=1)
            end_of_year = today.replace(month=12, day=31)
            paid_customers = paid_customers.filter(payment_date__range=[start_of_year, end_of_year])

        paid_customers = paid_customers.order_by(order_by)

        return paid_customers

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["payment_status"] = self.request.GET.get("payment_status", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")

        context["contacts"] = Contact.objects.all()

        if self.request.user.is_authenticated:
            if self.request.user.role == 2:
                # Get the current team leader
                team_leader = TeamLeader.objects.get(auth_user=self.request.user)
                # Get the agents under the current team leader
                agents = Agent.objects.filter(team_leader=team_leader)
                agent_user_ids = agents.values_list('auth_user', flat=True)
                # Get the matching CustomUser objects for the agent_user_ids
                leads = Lead.objects.filter(Q(assigned_to=team_leader.id) | Q(assigned_to_agents__in=agent_user_ids))
                context["leads"] = leads
            elif self.request.user.role == 3:
                agent = Agent.objects.get(auth_user=self.request.user)
                leads = Lead.objects.filter(
                    Q(assigned_to_agents=agent) | Q(assigned_to=agent.team_leader)
                )
                context["leads"] = leads
            else:
                context['leads'] = Lead.objects.all()

        if self.request.user.is_authenticated:
            current_user = self.request.user
            PaidCustomerNotification.objects.filter(user=current_user).update(seen=True)

        return context


class PaidCustomerCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = PaidCustomer
    fields = ['payment_date', 'related_contact', 'transaction_id', 'amount_paid', 'payment_status', 'payment_method', 'lead', 'remarks']
    template_name = 'paidcustomer_list.html'
    success_url = reverse_lazy('paidcustomer_list')
    success_message = "Paid Customer created!"

    def form_valid(self, form):
        self.object = form.save()
        return redirect(self.get_success_url())
    
    def form_invalid(self, form):
        for field in form.errors:
            error_messages = form.errors[field]
            for error_message in error_messages:
                messages.error(self.request, f"{field.capitalize()} : {error_message}")
        return HttpResponseRedirect(reverse_lazy('paidcustomer_list'))


class PaidCustomerUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = PaidCustomer
    fields = ['payment_date', 'related_contact', 'transaction_id', 'amount_paid', 'payment_status', 'payment_method', 'lead', 'remarks']
    template_name = 'paidcustomer_update.html'
    success_url = reverse_lazy('paidcustomer_list')
    success_message = "Paid Customer updated!"

    def get_form(self, form_class=None):
            form = super().get_form(form_class=form_class)
            current_user = self.request.user
            if self.request.user.is_authenticated:
                if current_user.role == 2:  # Team Leader
                    try:
                        team_leader_instance = TeamLeader.objects.get(auth_user=current_user)
                        agent_ids = Agent.objects.filter(team_leader=team_leader_instance).values_list('auth_user', flat=True)
                        form.fields['lead'].queryset = Lead.objects.filter(
                            Q(assigned_to=current_user.teamleader.id) | Q(assigned_to_agents__auth_user__in=agent_ids)
                        )
                    except TeamLeader.DoesNotExist:
                        form.fields['lead'].queryset = Lead.objects.none()
                elif current_user.role == 3:  # Agent
                    agent = get_object_or_404(Agent, auth_user=current_user)
                    form.fields['lead'].queryset = Lead.objects.filter(
                        Q(assigned_to=agent.team_leader.id) | Q(assigned_to_agents=agent.id)
                    )

                if current_user.role == 2 or current_user.role == 3:  # Team Leader or Agent
               
                    form.fields['payment_status'].choices = [('Pending', 'Pending')]
            return form



class PaidCustomerDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = PaidCustomer
    template_name = 'paidcustomer_list.html'
    success_url = reverse_lazy('paidcustomer_list')
    success_message = "Paid Customer deleted!"

    def test_func(self):
        # Check if the user has permission to delete the paid customer
        paid_customer = self.get_object()
        # You can modify this logic based on your requirements
        return self.request.user.role == 2  # O



class ComplaintListView(LoginRequiredMixin, ListView):
    model = Complaint
    template_name = 'complaint_list.html'
    paginate_by = 10

    def get_queryset(self):
        filter_val = self.request.GET.get("filter", "")
        resolved = self.request.GET.get("resolved", "")
        date_from = self.request.GET.get("date_from", "")
        date_to = self.request.GET.get("date_to", "")
        order_by = self.request.GET.get("orderby", "-id")

        complaints = Complaint.objects.all()

        if filter_val:
            complaints = complaints.filter(
                Q(related_contact__name__icontains=filter_val) |
                Q(related_contact__phone_number__icontains=filter_val) |
                Q(complaint_text__icontains=filter_val)
            )

        if resolved:
            resolved = resolved.lower() == "true"
            complaints = complaints.filter(resolved=resolved)

        if date_from and date_to:
            # Convert date strings to datetime objects
            date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            date_to = datetime.strptime(date_to, "%Y-%m-%d").date()
            # Add one day to the end date to include calls on that day
            date_to += timedelta(days=1)
            complaints = complaints.filter(created_at__range=(date_from, date_to))

        if self.request.user.is_authenticated:
            if self.request.user.role == 2:  # Team Leader
                team_leader = TeamLeader.objects.get(auth_user=self.request.user)
                # Retrieve the agents associated with the current Team Leader
                agents = Agent.objects.filter(team_leader=team_leader)
                # Retrieve the associated CustomUser instances for agents
                agent_users = agents.values_list('auth_user', flat=True)
                # Include complaints for both agents and the current Team Leader
                complaints = complaints.filter(Q(assigned_to__in=agent_users) | Q(assigned_to=self.request.user))
            elif self.request.user.role == 3:  # Agent
                complaints = complaints.filter(assigned_to=self.request.user)

        filter_by = self.request.GET.get("filter_by")
        if filter_by == "today":
            today = timezone.now().date()
            complaints = complaints.filter(created_at=today)
        elif filter_by == "this_week":
            start_of_week = timezone.now().date() - timezone.timedelta(days=timezone.now().weekday())
            end_of_week = start_of_week + timezone.timedelta(days=6)
            complaints = complaints.filter(created_at__range=[start_of_week, end_of_week])
        elif filter_by == "this_month":
            today = timezone.now().date()
            start_of_month = today.replace(day=1)
            end_of_month = start_of_month.replace(day=calendar.monthrange(start_of_month.year, start_of_month.month)[1])
            complaints = complaints.filter(created_at__range=[start_of_month, end_of_month])
        elif filter_by == "6_months":
            today = timezone.now().date()
            start_of_six_months_ago = today - timezone.timedelta(days=180)
            complaints = complaints.filter(created_at__gte=start_of_six_months_ago)
        elif filter_by == "this_year":
            today = timezone.now().date()
            start_of_year = today.replace(month=1, day=1)
            end_of_year = today.replace(month=12, day=31)
            complaints = complaints.filter(created_at__range=[start_of_year, end_of_year])

        
        complaints = complaints.order_by(order_by)

        return complaints

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["status"] = self.request.GET.get("status", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")
        context['contacts'] = Contact.objects.all()

        users = CustomUser.objects.exclude(id=self.request.user.id)

        if self.request.user.is_authenticated:
            if self.request.user.role == 2:
                # Get the current team leader
                team_leader = TeamLeader.objects.get(auth_user=self.request.user)
                # Get the agents under the current team leader
                agents = Agent.objects.filter(team_leader=team_leader)
                agent_user_ids = agents.values_list('auth_user', flat=True)
                # Get the matching CustomUser objects for the agent_user_ids
                users = CustomUser.objects.filter(Q(id=self.request.user.id) | Q(id__in=agent_user_ids))
                context["users"] = users
                # Update the agent information for agents that are under the current team leader
                for user in context["users"]:
                    user.team_leader_first_name = team_leader.first_name
                    user.team_leader_last_name = team_leader.last_name
                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user.agent_first_name = agent.first_name
                        user.agent_last_name = agent.last_name
                    except Agent.DoesNotExist:
                        user.agent_first_name = None
                        user.agent_last_name = None

            elif self.request.user.role == 3:
                users = CustomUser.objects.filter(id=self.request.user.id)
                context["users"] = users

                for user in context["users"]:
                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user.agent_first_name = agent.first_name
                        user.agent_last_name = agent.last_name
                    except Agent.DoesNotExist:
                        user.agent_first_name = None
                        user.agent_last_name = None
            else:
                for user in users:
                    try:
                        team_leader = TeamLeader.objects.get(auth_user=user)
                        user.team_leader_first_name = team_leader.first_name
                        user.team_leader_last_name = team_leader.last_name
                    except TeamLeader.DoesNotExist:
                        user.team_leader_first_name = None
                        user.team_leader_last_name = None

                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user.agent_first_name = agent.first_name
                        user.agent_last_name = agent.last_name
                    except Agent.DoesNotExist:
                        user.agent_first_name = None
                        user.agent_last_name = None

                context["users"] = users

        if self.request.user.is_authenticated:
            current_user = self.request.user
            ComplaintNotification.objects.filter(user=current_user).update(seen=True)

        return context


class ComplaintCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Complaint
    fields = ['assigned_to', 'related_contact', 'complaint_text', 'resolved']
    template_name = 'complaint_list.html'
    success_url = reverse_lazy('complaint_list')
    success_message = "Complaint Created!"

    def form_valid(self, form):
        self.object = form.save()
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        for field in form.errors:
            error_messages = form.errors[field]
            for error_message in error_messages:
                messages.error(self.request, f"{field.capitalize()} : {error_message}")
        return HttpResponseRedirect(reverse_lazy('complaint_list'))


class ComplaintUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Complaint
    fields = ['assigned_to', 'related_contact', 'complaint_text', 'resolved']
    template_name = 'complaint_update.html'
    success_url = reverse_lazy('complaint_list')
    success_message = "Complaint Updated!"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if self.request.user.role == 2:  # Team Leader
            team_leader = get_object_or_404(TeamLeader, auth_user=self.request.user)
            agents = Agent.objects.filter(team_leader=team_leader)
            # Retrieve the associated CustomUser instances for agents
            agent_users = agents.values_list('auth_user', flat=True)
            # Filter the 'assigned_to' field queryset to include only the agents
            form.fields['assigned_to'].queryset = CustomUser.objects.filter(id__in=agent_users)
        elif self.request.user.role == 3:  # Agent
            # Set the 'assigned_to' field queryset to include only the current user agent
            agent = get_object_or_404(Agent, auth_user=self.request.user)
            form.fields['assigned_to'].queryset = CustomUser.objects.filter(id=agent.auth_user.id)
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.object and self.object.assigned_to:
            # Assuming that the 'assigned_to' field is a ForeignKey to 'CustomUser'
            agent = self.object.assigned_to
            context['agent_first_name'] = agent.first_name
            context['agent_last_name'] = agent.last_name
        return context


class ComplaintDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Complaint
    template_name = 'complaint_list.html'
    success_url = reverse_lazy('complaint_list')
    success_message = "Complaint Deleted!"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        return redirect(self.get_success_url())


class TaskListView(LoginRequiredMixin, ListView):
    model = Task
    template_name = 'task_list.html'
    paginate_by = 10

    def get_queryset(self):
        filter_val = self.request.GET.get("filter", "")
        status = self.request.GET.get("status", "")
        due_date_from = self.request.GET.get("due_date_from", "")
        due_date_to = self.request.GET.get("due_date_to", "")
        order_by = self.request.GET.get("orderby", "-id")

        tasks = Task.objects.all()

        if filter_val:
            tasks = tasks.filter(
                Q(name__icontains=filter_val) |
                Q(description__icontains=filter_val) |
                Q(date__icontains=filter_val) |
                Q(status__icontains=filter_val)
            )

        if status:
            tasks = tasks.filter(status=status)

        if due_date_from and due_date_to:
            tasks = tasks.filter(due_date__range=(due_date_from, due_date_to))

        # Add any additional filtering or querying logic here
        if self.request.user.is_authenticated and self.request.user.role == 2:
            team_leader = TeamLeader.objects.get(auth_user=self.request.user)
            agent_ids = Agent.objects.filter(team_leader=team_leader).values_list('auth_user', flat=True)
            tasks = tasks.filter(Q(assigned_to=self.request.user) | Q(assigned_to__in=agent_ids))
            
        elif self.request.user.is_authenticated and self.request.user.role == 3:
       
            tasks = tasks.filter(assigned_to=self.request.user)
        

        tasks = tasks.order_by(order_by)

        return tasks

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["status"] = self.request.GET.get("status", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")
        context['contacts'] = Contact.objects.all()

        users = CustomUser.objects.exclude(id=self.request.user.id)

        if self.request.user.is_authenticated:
            if self.request.user.role == 2:
                # Get the current team leader
                team_leader = TeamLeader.objects.get(auth_user=self.request.user)
                # Get the agents under the current team leader
                agents = Agent.objects.filter(team_leader=team_leader)
                agent_user_ids = agents.values_list('auth_user', flat=True)
                # Get the matching CustomUser objects for the agent_user_ids
                users = CustomUser.objects.filter(Q(id=self.request.user.id) | Q(id__in=agent_user_ids))
                context["users"] = users
                # Update the agent information for agents that are under the current team leader
                for user in context["users"]:
                    user.team_leader_first_name = team_leader.first_name
                    user.team_leader_last_name = team_leader.last_name
                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user.agent_first_name = agent.first_name
                        user.agent_last_name = agent.last_name
                    except Agent.DoesNotExist:
                        user.agent_first_name = None
                        user.agent_last_name = None

            elif self.request.user.role == 3:
                
                users = CustomUser.objects.filter(id=self.request.user.id)
                context["users"] = users

                for user in context["users"]:
                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user.agent_first_name = agent.first_name
                        user.agent_last_name = agent.last_name
                    except Agent.DoesNotExist:
                        user.agent_first_name = None
                        user.agent_last_name = None

            else:
                for user in users:
                    try:
                        team_leader = TeamLeader.objects.get(auth_user=user)
                        user.team_leader_first_name = team_leader.first_name
                        user.team_leader_last_name = team_leader.last_name
                    except TeamLeader.DoesNotExist:
                        user.team_leader_first_name = None
                        user.team_leader_last_name = None

                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user.agent_first_name = agent.first_name
                        user.agent_last_name = agent.last_name
                    except Agent.DoesNotExist:
                        user.agent_first_name = None
                        user.agent_last_name = None

                context["users"] = users

        if self.request.user.is_authenticated:
            current_user = self.request.user
            TaskNotification.objects.filter(user=current_user).update(seen=True)

        return context

class TaskCreateView(SuccessMessageMixin, CreateView):
    model = Task
    fields = ['date', 'name', 'task_type', 'due_date', 'priority', 'status', 'description', 'related_contact', 'assigned_to']
    template_name = 'task_list.html'
    success_url = reverse_lazy('task_list')
    success_message = "Task Created!"

    def form_valid(self, form):

        if self.request.user.is_authenticated and self.request.user.role == 3:
            agent = get_object_or_404(Agent, auth_user=self.request.user)
            form.instance.assigned_to = agent

        self.object = form.save()
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        for field in form.errors:
            error_messages = form.errors[field]
            for error_message in error_messages:
                messages.error(self.request, f"{field.capitalize()} : {error_message}")
        return HttpResponseRedirect(reverse_lazy('task_list'))


class TaskUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Task
    fields = ['date', 'name', 'task_type', 'due_date', 'priority', 'status', 'description', 'related_contact', 'assigned_to']
    template_name = 'task_update.html'
    success_url = reverse_lazy('task_list')
    success_message = "Task Updated!"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if self.request.user.role == 2:  # Team Leader
            team_leader = get_object_or_404(TeamLeader, auth_user=self.request.user)
            agents = Agent.objects.filter(team_leader=team_leader)
            # Retrieve the associated CustomUser instances for agents
            agent_users = agents.values_list('auth_user', flat=True)
            # Filter the 'assigned_to' field queryset to include only the agents
            form.fields['assigned_to'].queryset = CustomUser.objects.filter(id__in=agent_users)
        elif self.request.user.role == 3:  # Agent
            # Set the 'assigned_to' field queryset to include only the current user agent
            agent = get_object_or_404(Agent, auth_user=self.request.user)
            form.fields['assigned_to'].queryset = CustomUser.objects.filter(id=agent.auth_user.id)
        return form

class TaskDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Task
    template_name = 'task_list.html'
    success_url = reverse_lazy('task_list')
    success_message = "Task Deleted!"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        return redirect(self.get_success_url())


class MarketingListView(LoginRequiredMixin, ListView):
    model = Marketing
    template_name = 'marketing_list.html'
    paginate_by = 10

    def get_queryset(self):
        filter_val = self.request.GET.get("filter", "")
        campaign_type = self.request.GET.get("campaign_type", "")
        campaign_objective = self.request.GET.get("campaign_objective", "")
        campaign_status = self.request.GET.get("campaign_status", "")
        date_from = self.request.GET.get("date_from", "")
        date_to = self.request.GET.get("date_to", "")
        order_by = self.request.GET.get("orderby", "-id")

        campaigns = Marketing.objects.all()

        if filter_val:
            campaigns = campaigns.filter(
                Q(id__icontains=filter_val) |
                Q(date__icontains=filter_val) |
                Q(name__icontains=filter_val)
            )

        if campaign_type:
            campaigns = campaigns.filter(campaign_type=campaign_type)

        if campaign_objective:
            campaigns = campaigns.filter(campaign_objective=campaign_objective)

        if campaign_status:
            campaigns = campaigns.filter(campaign_status=campaign_status)

        if date_from and date_to:
            # Convert the date strings to datetime objects
            date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            date_to = datetime.strptime(date_to, "%Y-%m-%d").date()

            campaigns = campaigns.filter(date__range=(date_from, date_to))

        campaigns = campaigns.order_by(order_by)

        return campaigns

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["campaign_type"] = self.request.GET.get("campaign_type", "")
        context["campaign_objective"] = self.request.GET.get(
            "campaign_objective", "")
        context["campaign_status"] = self.request.GET.get(
            "campaign_status", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")

        if self.request.user.is_authenticated:
            current_user = self.request.user
            MarketingNotification.objects.filter(user=current_user).update(seen=True)

        return context


class MarketingCreateView(LoginRequiredMixin, CreateView):
    model = Marketing
    fields = ['date', 'name', 'campaign_type', 'campaign_objective',
              'campaign_status', 'target_audience', 'campaign_budget', 'campaign_channels']
    template_name = 'marketing_list.html'
    success_url = reverse_lazy('marketing_list')
    success_message = "Marketing Campaign Created!"

    def form_valid(self, form):
        self.object = form.save()
        return redirect(self.get_success_url())
    
    def form_invalid(self, form):
        for field in form.errors:
            error_messages = form.errors[field]
            for error_message in error_messages:
                messages.error(self.request, f"{field.capitalize()} : {error_message}")
        return HttpResponseRedirect(reverse_lazy('marketing_list'))


class MarketingUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Marketing
    fields = ['date', 'name', 'campaign_type', 'campaign_objective',
              'campaign_status', 'target_audience', 'campaign_budget', 'campaign_channels']
    template_name = 'marketing_update.html'
    success_url = reverse_lazy('marketing_list')
    success_message = "Marketing Campaign Updated!"


class MarketingDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Marketing
    template_name = 'marketing_list.html'
    success_url = reverse_lazy('marketing_list')
    success_message = "Marketing Campaign Deleted!"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        return redirect(self.get_success_url())


class CalendarListView(LoginRequiredMixin, ListView):
    model = Calendar
    template_name = 'calendar_list.html'
    paginate_by = 10

    def get_queryset(self):
        filter_val = self.request.GET.get("filter", "")
        event_date = self.request.GET.get("event_date", "")
        location = self.request.GET.get("location", "")
        participants = self.request.GET.get("participants", "")
        date_from = self.request.GET.get("date_from", "")
        date_to = self.request.GET.get("date_to", "")
        order_by = self.request.GET.get("orderby", "-id")

        calendars = Calendar.objects.all()

        if filter_val:
            calendars = calendars.filter(
                Q(name__icontains=filter_val) |
                Q(event_date__icontains=filter_val) |
                Q(participants__icontains=filter_val) |
                Q(location__icontains=filter_val)
            )

        if date_from and date_to:
            # Convert the date strings to datetime objects
            date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            date_to = datetime.strptime(date_to, "%Y-%m-%d").date()

            calendars = calendars.filter(
                event_date__range=(date_from, date_to))
            
        # Add any additional filtering or querying logic here
        if self.request.user.is_authenticated and self.request.user.role == 2:
            team_leader = TeamLeader.objects.get(auth_user=self.request.user)
            agent_ids = Agent.objects.filter(team_leader=team_leader).values_list('auth_user', flat=True)
            calendars = calendars.filter(Q(assigned_to=self.request.user) | Q(assigned_to__in=agent_ids))
        elif self.request.user.is_authenticated and self.request.user.role == 3:
            calendars = calendars.filter(assigned_to=self.request.user)

        filter_by = self.request.GET.get("filter_by")
        if filter_by == "today":
            today = timezone.now().date()
            calendars = calendars.filter(event_date=today)
        elif filter_by == "this_week":
            start_of_week = timezone.now().date() - timezone.timedelta(days=timezone.now().weekday())
            end_of_week = start_of_week + timezone.timedelta(days=6)
            calendars = calendars.filter(event_date__range=[start_of_week, end_of_week])
        elif filter_by == "this_month":
            today = timezone.now().date()
            start_of_month = today.replace(day=1)
            end_of_month = start_of_month.replace(day=calendar.monthrange(start_of_month.year, start_of_month.month)[1])
            calendars = calendars.filter(event_date__range=[start_of_month, end_of_month])
        elif filter_by == "6_months":
            today = timezone.now().date()
            start_of_six_months_ago = today - timezone.timedelta(days=180)
            calendars = calendars.filter(event_date__gte=start_of_six_months_ago)
        elif filter_by == "this_year":
            today = timezone.now().date()
            start_of_year = today.replace(month=1, day=1)
            end_of_year = today.replace(month=12, day=31)
            calendars = calendars.filter(event_date__range=[start_of_year, end_of_year])

        calendars = calendars.order_by(order_by)

        return calendars

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter"] = self.request.GET.get("filter", "")
        context["event_date"] = self.request.GET.get("event_date", "")
        context["location"] = self.request.GET.get("location", "")
        context["participants"] = self.request.GET.get("participants", "")
        context["orderby"] = self.request.GET.get("orderby", "-id")

        users = CustomUser.objects.exclude(id=self.request.user.id)

        if self.request.user.is_authenticated:
            if self.request.user.role == 2:
                # Get the current team leader
                team_leader = TeamLeader.objects.get(auth_user=self.request.user)
                # Get the agents under the current team leader
                agents = Agent.objects.filter(team_leader=team_leader)
                agent_user_ids = agents.values_list('auth_user', flat=True)
                # Get the matching CustomUser objects for the agent_user_ids
                users = CustomUser.objects.filter(Q(id=self.request.user.id) | Q(id__in=agent_user_ids))
                context["users"] = users

                for user in context["users"]:
                    user.team_leader_first_name = team_leader.first_name
                    user.team_leader_last_name = team_leader.last_name
                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user.agent_first_name = agent.first_name
                        user.agent_last_name = agent.last_name
                    except Agent.DoesNotExist:
                        user.agent_first_name = None
                        user.agent_last_name = None
            elif self.request.user.role == 3:
                users = CustomUser.objects.filter(id=self.request.user.id)
                context["users"] = users

                for user in context["users"]:
                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user.agent_first_name = agent.first_name
                        user.agent_last_name = agent.last_name
                    except Agent.DoesNotExist:
                        user.agent_first_name = None
                        user.agent_last_name = None            
            else:
                for user in users:
                    try:
                        team_leader = TeamLeader.objects.get(auth_user=user)
                        user.team_leader_first_name = team_leader.first_name
                        user.team_leader_last_name = team_leader.last_name
                    except TeamLeader.DoesNotExist:
                        user.team_leader_first_name = None
                        user.team_leader_last_name = None

                    try:
                        agent = Agent.objects.get(auth_user=user)
                        user.agent_first_name = agent.first_name
                        user.agent_last_name = agent.last_name
                    except Agent.DoesNotExist:
                        user.agent_first_name = None
                        user.agent_last_name = None

                context["users"] = users

        context['contacts'] = Contact.objects.all()

        if self.request.user.is_authenticated:
            current_user = self.request.user
            CalendarNotification.objects.filter(user=current_user).update(seen=True)

        return context


class CalendarCreateView(LoginRequiredMixin, CreateView):
    model = Calendar
    fields = ['name', 'event_date', 'location', 'participants', 'description',
              'reminders', 'related_contact', 'assigned_to']
    template_name = 'calendar_list.html'
    success_url = reverse_lazy('calendar_list')
    success_message = "Lead Created!"

    def form_valid(self, form):
        self.object = form.save()
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        for field in form.errors:
            error_messages = form.errors[field]
            for error_message in error_messages:
                messages.error(self.request, f"{field.capitalize()} : {error_message}")
        return HttpResponseRedirect(reverse_lazy('calendar_list'))


class CalendarUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Calendar
    fields = ['name', 'event_date', 'location', 'participants', 'description',
              'reminders', 'related_contact', 'assigned_to']
    template_name = 'calendar_update.html'
    success_url = reverse_lazy('calendar_list')
    success_message = "Calendar Updated!"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if self.request.user.role == 2:  # Team Leader
            team_leader = get_object_or_404(TeamLeader, auth_user=self.request.user)
            agents = Agent.objects.filter(team_leader=team_leader)
           
            custom_users = agents.values_list('auth_user', flat=True)

            # Update the queryset for assigned_to field with the CustomUser instances
            form.fields['assigned_to'].queryset = CustomUser.objects.filter(pk__in=custom_users)
        elif self.request.user.role == 3:  # Agent
            # Set the 'assigned_to' field queryset to include only the current user agent
            agent = get_object_or_404(Agent, auth_user=self.request.user)
            form.fields['assigned_to'].queryset = CustomUser.objects.filter(id=agent.auth_user.id)
        return form
       


class CalendarDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Calendar
    template_name = 'calendar_list.html'
    success_url = reverse_lazy('calendar_list')
    success_message = "Calendar Deleted!"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        return redirect(self.get_success_url())


def lead_analysis(request):
    # Lead Status Distribution
    status_counts = Lead.objects.values('status').annotate(count=Count('status'))

    # Source Analysis
    source_counts = Lead.objects.values('source').annotate(count=Count('source'))

    # Follow-up Action Analysis
    action_counts = Lead.objects.values('follow_up_actions').annotate(count=Count('follow_up_actions'))

    # Lead Assignment Analysis
    assignment_counts = Lead.objects.values('assigned_to__username').annotate(count=Count('assigned_to'))

    # Time-Based Analysis
    time_counts = Lead.objects.annotate(date_str=Cast('date', output_field=CharField())).values('date_str').annotate(count=Count('date'))

    # Convert QuerySet data to list of dictionaries
    status_data = list(status_counts.values('status', 'count'))
    source_data = list(source_counts.values('source', 'count'))
    action_data = list(action_counts.values('follow_up_actions', 'count'))
    assignment_data = list(assignment_counts.values('assigned_to__username', 'count'))
    time_data = list(time_counts.values('date_str', 'count'))

    # Serialize data to JSON format
    status_data = json.dumps(status_data)
    source_data = json.dumps(source_data)
    action_data = json.dumps(action_data)
    assignment_data = json.dumps(assignment_data)
    time_data = json.dumps(time_data)

    return render(request, 'leads_analysis.html', {
        'status_data': status_data,
        'source_data': source_data,
        'action_data': action_data,
        'assignment_data': assignment_data,
        'time_data': time_data,
    })


def analysis(request):
    # Employee Statistics
    total_team_leaders = TeamLeader.objects.count()
    total_agents = Agent.objects.count()
    total_employees  = total_team_leaders + total_agents

    # Lead Statistics
    total_leads = Lead.objects.count()
    leads_by_status_new = Lead.objects.filter(status='new').count()
    leads_by_status_contacted = Lead.objects.filter(status='contacted').count()
    leads_by_status_converted = Lead.objects.filter(status='converted').count()
    leads_by_status_lost = Lead.objects.filter(status='lost').count()

    # Paid Customer Statistics
    total_paid_customers = PaidCustomer.objects.count()
    paid_customers_by_status_pending = PaidCustomer.objects.filter(payment_status='pending').count()
    paid_customers_by_status_completed = PaidCustomer.objects.filter(payment_status='completed').count()
    paid_customers_by_status_refunded = PaidCustomer.objects.filter(payment_status='refunded').count()
    paid_customers_by_status_cancelled = PaidCustomer.objects.filter(payment_status='cancelled').count()

    # Task Statistics
    total_tasks = Task.objects.count()
    tasks_by_status_pending = Task.objects.filter(status='pending').count()
    tasks_by_status_in_progress = Task.objects.filter(status='in_progress').count()
    tasks_by_status_completed = Task.objects.filter(status='completed').count()
    tasks_by_status_cancelled = Task.objects.filter(status='cancelled').count()

    # Complaint Statistics
    total_complaints = Complaint.objects.count()
    resolved_complaints = Complaint.objects.filter(resolved=True).count()
    unresolved_complaints = Complaint.objects.filter(resolved=False).count()
  
    # Marketing Campaign Statistics
    total_campaigns = Marketing.objects.count()
    campaigns_by_status_active = Marketing.objects.filter(campaign_status='active').count()
    campaigns_by_status_paused = Marketing.objects.filter(campaign_status='paused').count()
    campaigns_by_status_completed = Marketing.objects.filter(campaign_status='completed').count()

    total_contacts = Contact.objects.count()
    total_leads_converted = Lead.objects.filter(status='converted').count()
    total_tasks_completed = Task.objects.filter(status='completed').count()
    total_complaints_resolved = Complaint.objects.filter(resolved=True).count()
    total_marketing_completed = Marketing.objects.filter(campaign_status='completed').count()
    total_paid_customers_completed = PaidCustomer.objects.filter(payment_status='completed').count()

    chart_data = {
        'labels': ['Contacts', 'Leads(Converted)', 'Paid Customers(Completed)', 'Tasks(Completed)', 
                   'Complaints(Resolved)', 'Marketing(Completed)'],
        'data': [total_contacts, total_leads_converted,total_paid_customers_completed, total_tasks_completed, 
                 total_complaints_resolved, total_marketing_completed],
    }

    context = {
        # Employee Statistics
        'total_team_leaders': total_team_leaders,
        'total_agents': total_agents,
        'total_employees':total_employees,

        # Lead Statistics
        'total_leads': total_leads,
        'leads_by_status_new': leads_by_status_new,
        'leads_by_status_contacted': leads_by_status_contacted,
        'leads_by_status_converted': leads_by_status_converted,
        'leads_by_status_lost': leads_by_status_lost,


        'total_paid_customers': total_paid_customers,
        'paid_customers_by_status_pending': paid_customers_by_status_pending,
        'paid_customers_by_status_completed': paid_customers_by_status_completed,
        'paid_customers_by_status_refunded': paid_customers_by_status_refunded,
        'paid_customers_by_status_cancelled': paid_customers_by_status_cancelled,
        

        # Task Statistics
        'total_tasks': total_tasks,
        'tasks_by_status_pending': tasks_by_status_pending,
        'tasks_by_status_completed': tasks_by_status_completed,
        'tasks_by_status_in_progress': tasks_by_status_in_progress,
        'tasks_by_status_cancelled': tasks_by_status_cancelled,
      

        # Complaint Statistics
        'total_complaints': total_complaints,
        'resolved_complaints': resolved_complaints,
        'unresolved_complaints': unresolved_complaints,

        # Marketing Campaign Statistics
        'total_campaigns': total_campaigns,
        'campaigns_by_status_active': campaigns_by_status_active,
        'campaigns_by_status_paused': campaigns_by_status_paused,
        'campaigns_by_status_completed': campaigns_by_status_completed,
       
        # Charts Data
        'chart_data_json': json.dumps(chart_data),
        
    }

    return render(request, 'overall_analysis.html', context)

from django.shortcuts import render
from django.db.models import Sum, F, ExpressionWrapper, FloatField, Q
from .models import Agent, AgentSales, AgentAttendance, PaidCustomer, Lead




def agent_sales_view(request):

    filter_param = request.GET.get('filter', None)
    orderby = request.GET.get('orderby', None)

    if request.user.role == 2:
        user = request.user
        current_user_team_leader_id = TeamLeader.objects.get(auth_user=user)

        # Filter agents belonging to the current user's team
        agents = Agent.objects.filter(team_leader_id=current_user_team_leader_id)
    else:
        # If the current user is not a team leader, show all agents
        agents = Agent.objects.all()

    for agent in agents:
        sales_amount = PaidCustomer.objects.filter(lead__assigned_to_agents=agent).aggregate(total_sales=Sum('amount_paid'))['total_sales'] or 0.0
        total_days = AgentAttendance.objects.filter(agent=agent).count()
        present_days = AgentAttendance.objects.filter(agent=agent).filter(Q(status='Present') | Q(status='Late')).count()
        attendance_percentage = (present_days / total_days) * 100 if total_days > 0 else 0.0
        count_leads = Lead.objects.filter(assigned_to_agents=agent).count()
        lead_conversion = (count_leads / sales_amount) * 100 if sales_amount > 0 else 0.0
        lead_completed = Lead.objects.filter(assigned_to_agents=agent, status='converted').count()
        paid_customers_count = PaidCustomer.objects.filter(lead__assigned_to_agents=agent, payment_status='completed').count()
        sales_achievement = (paid_customers_count / 10) * 100
        team_leader = agent.team_leader

        # Use update_or_create to either update the existing AgentSales object or create a new one
        agent_sales, created = AgentSales.objects.update_or_create(
            agent=agent,
            defaults={
                'team_leader': team_leader,
                'sale_amount': sales_amount,
                'attendance_percentage': attendance_percentage,
                'count_leads': count_leads,
                'lead_conversion': lead_conversion,
                'lead_completed': lead_completed,
                'sales_achievement': sales_achievement,
            }
        )


    agent_sales_data = AgentSales.objects.order_by('-sale_amount')

    if filter_param:
        agent_sales_data = agent_sales_data.filter(
            Q(agent__first_name__icontains=filter_param) |
            Q(agent__last_name__icontains=filter_param)
        )

    # Apply sorting based on orderby parameter
    if orderby == 'sale_amount':
        agent_sales_data = agent_sales_data.order_by('-sale_amount')
    elif orderby == 'attendance_percentage':
        agent_sales_data = agent_sales_data.order_by('-attendance_percentage')
    
    return render(request, 'agent_sales.html', {
        'agent_sales_data': agent_sales_data,
        'filter_param': filter_param,
        'orderby': orderby,
    })


def agent_update_target(request, agent_id):
    agent = Agent.objects.get(pk=agent_id)

    if request.method == 'POST':
        new_sales_target = request.POST.get('sales_target')
        agent.sales_traget = new_sales_target
        agent.save()
        return redirect('agent_list') 

def team_leader_update_target(request, team_leader_id):
    team_leader = TeamLeader.objects.get(pk=team_leader_id)

    if request.method == 'POST':
        new_sales_target = request.POST.get('sales_target')
        team_leader.sales_traget = new_sales_target
        team_leader.save()
        return redirect('team_leader_list') 

















